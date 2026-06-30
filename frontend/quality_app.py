"""
Streamlit dashboard for the Openhouse quality team.

Run:
    .venv/bin/streamlit run frontend/quality_app.py

Reads the existing pipeline artefacts directly — no DB:
  data/tqm_answers.xlsx     ← rolling Q&A log (one row per Q × run)
  data/rubric_runs/<subj>/<config_slug>/5_answers.json     (full payload)
  data/rubric_runs/<subj>/<config_slug>/0_config.json      (activity_context)
  data/sessions/<subj>/<sid>/3_trimmed.mp4                  (player source)
  data/cctv_cameras.xlsx                                    (camera → centre)

Two-pane layout:
  LEFT  — sessions list (one row per session — the "canonical" run only).
  RIGHT — selected session: video at top, materials + class window beside it,
          then four tabs (Environment / Content Knowledge / Facilitation /
          Warmth). Click an evidence timestamp anywhere → video seeks.
"""
from __future__ import annotations

import json
import os
from datetime import date, datetime, time, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st

ROOT = Path(__file__).resolve().parent.parent
RUBRIC_RUNS = ROOT / "data" / "rubric_runs"
SESSIONS_DIR = ROOT / "data" / "sessions"
CAMERAS_XLSX = ROOT / "data" / "cctv_cameras.xlsx"

# Subject tokens — same set the pipeline uses end-to-end.
SUBJECTS = ("art", "public_speaking", "robotics")

# 5-min margin on each side of the class window (the convention I use when
# manually trimming via the fast path). Used only to estimate class end time.
TRIM_MARGIN_MIN = 5

SECTION_TABS = ("Environment", "Content Knowledge", "Facilitation", "Warmth")
SECTION_ICONS = {
    "Environment": "🏛️",
    "Content Knowledge": "📚",
    "Facilitation": "🤝",
    "Warmth": "💛",
}

st.set_page_config(
    page_title="Openhouse",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ─── Cloud detection + shared-password gate ──────────────────────────────────


def is_cloud_deploy() -> bool:
    """True when running on Streamlit Community Cloud (or any deploy where
    `is_cloud_deploy = true` is set in st.secrets). False when running
    locally without secrets."""
    try:
        return bool(st.secrets.get("is_cloud_deploy", False))
    except Exception:
        return False


def _r2_secrets() -> dict | None:
    """Return {account_id, bucket, key, secret} when all four R2 secrets are
    set in st.secrets (either local .streamlit/secrets.toml or the Streamlit
    Cloud Settings → Secrets panel). Returns None otherwise."""
    try:
        s = st.secrets
        needed = ("r2_account_id", "r2_bucket", "r2_access_key_id", "r2_secret_access_key")
        if not all(s.get(k) for k in needed):
            return None
        return {
            "account_id": s["r2_account_id"],
            "bucket": s["r2_bucket"],
            "key": s["r2_access_key_id"],
            "secret": s["r2_secret_access_key"],
        }
    except Exception:
        return None


@st.cache_resource
def _r2_client():
    """Construct (and cache for the lifetime of the Streamlit session) the
    boto3 S3 client pointing at the R2 endpoint. Returns None when R2
    secrets aren't configured."""
    creds = _r2_secrets()
    if creds is None:
        return None
    try:
        import boto3
        from botocore.config import Config
        return boto3.client(
            "s3",
            endpoint_url=f"https://{creds['account_id']}.r2.cloudflarestorage.com",
            aws_access_key_id=creds["key"],
            aws_secret_access_key=creds["secret"],
            region_name="auto",
            config=Config(signature_version="s3v4"),
        )
    except Exception:
        return None


@st.cache_data(ttl=1800)  # 30 min — well under the 1-hour URL TTL
def r2_signed_video_url(subject: str, session_id: str) -> str | None:
    """Return a 1-hour pre-signed GET URL for the session's video on R2.
    The key convention matches scripts/upload_videos_to_r2.py:
        <subject>/<session_id>.mp4
    Returns None when R2 isn't configured or the object can't be signed."""
    client = _r2_client()
    creds = _r2_secrets()
    if client is None or creds is None:
        return None
    key = f"{subject}/{session_id}.mp4"
    try:
        return client.generate_presigned_url(
            "get_object",
            Params={"Bucket": creds["bucket"], "Key": key},
            ExpiresIn=3600,
        )
    except Exception:
        return None


def password_gate() -> bool:
    """Shared-password lock used in cloud deploys (free tier has no built-in
    viewer auth). Returns True when the user is authenticated for this
    session; False otherwise (and renders the login UI as a side effect).

    Local runs without a `password` secret skip the gate entirely.
    """
    try:
        expected = st.secrets.get("password", None)
    except Exception:
        expected = None
    if expected is None:
        return True  # no secret configured → no gate (typical local dev)
    if st.session_state.get("_auth_ok"):
        return True

    with st.container():
        st.title("Openhouse")
        st.caption("Sign in to view the dashboard")
        entered = st.text_input("Password", type="password")
        if entered:
            if entered == expected:
                st.session_state["_auth_ok"] = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    return False


# ─── Data loaders ──────────────────────────────────────────────────────────


# ─── Supabase client ─────────────────────────────────────────────────────────


@st.cache_resource
def _supabase():
    """Cached Supabase client. Returns None when secrets aren't configured —
    callers should handle gracefully (empty data instead of crashing)."""
    try:
        url = st.secrets.get("SUPABASE_URL") or os.environ.get("SUPABASE_URL")
        key = (st.secrets.get("SUPABASE_SERVICE_KEY")
               or os.environ.get("SUPABASE_SERVICE_KEY"))
    except Exception:
        url = os.environ.get("SUPABASE_URL")
        key = os.environ.get("SUPABASE_SERVICE_KEY")
    if not url or not key:
        return None
    try:
        from supabase import create_client
        return create_client(url, key)
    except Exception:
        return None


def _fetch_all(table: str, *, select: str = "*", batch: int = 1000) -> list[dict]:
    """Paginate through every row of a Supabase table. The PostgREST API
    caps each response at 1000 rows by default, so we have to range-select."""
    sb = _supabase()
    if sb is None:
        return []
    out: list[dict] = []
    start = 0
    while True:
        end = start + batch - 1
        r = sb.table(table).select(select).range(start, end).execute()
        rows = r.data or []
        out.extend(rows)
        if len(rows) < batch:
            break
        start += batch
    return out


@st.cache_data(ttl=60)
def load_runs() -> pd.DataFrame:
    """One row per run, from the Supabase `runs` table."""
    rows = _fetch_all("runs")
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    if "session_date" in df.columns:
        df["session_date"] = pd.to_datetime(df["session_date"], errors="coerce").dt.date
    return df


@st.cache_data(ttl=60)
def runs_window_info() -> dict[str, dict]:
    """{run_id → {duration_sec, boundaries_detected}} — for class-window
    computation on the cloud build (where the video isn't on disk)."""
    df = load_runs()
    if df.empty:
        return {}
    out: dict[str, dict] = {}
    for _, row in df.iterrows():
        rid = str(row.get("run_id") or "").strip()
        if not rid:
            continue
        dur = row.get("trimmed_video_duration_seconds")
        bd = row.get("boundaries_detected")
        try:
            dur = float(dur) if pd.notna(dur) else None
        except Exception:
            dur = None
        try:
            bd_norm = bool(bd) if pd.notna(bd) else None
        except Exception:
            bd_norm = None
        out[rid] = {"duration_sec": dur, "boundaries_detected": bd_norm}
    return out


@st.cache_data(ttl=60)
def materials_by_run() -> dict[str, list]:
    """{run_id → parsed materials_seen list}. The runs.materials_seen column
    in Supabase is jsonb so we already get a Python list back."""
    df = load_runs()
    if df.empty:
        return {}
    out: dict[str, list] = {}
    for _, row in df.iterrows():
        rid = str(row.get("run_id") or "").strip()
        ms = row.get("materials_seen")
        if not rid or ms is None:
            continue
        # jsonb comes back as a list[dict] already; tolerate string-encoded too
        if isinstance(ms, str):
            try:
                ms = json.loads(ms)
            except Exception:
                continue
        if isinstance(ms, list) and ms:
            out[rid] = ms
    return out


@st.cache_data(ttl=60)
def load_all_answers() -> pd.DataFrame:
    """One row per (run × question) joined with run-level metadata the
    dashboard needs to filter (session_date, camera, teacher_id, run_n).
    Pulls from the Supabase `answers` + `runs` tables."""
    answers = _fetch_all("answers")
    if not answers:
        return pd.DataFrame()
    runs = load_runs()
    df = pd.DataFrame(answers)
    if not runs.empty:
        wanted = ["run_id", "session_date", "camera", "rubric_version",
                  "reasoner", "shape", "run_n", "vision_model"]
        keep = [c for c in wanted if c in runs.columns]
        df = df.merge(runs[keep], on="run_id", how="left")
    # The answers table has insufficient_information as a real bool — but
    # pandas may bring it back as Python bool or numpy bool depending on
    # the supabase-py version. Normalise.
    if "insufficient_information" in df.columns:
        df["insufficient_information"] = df["insufficient_information"].astype(bool)
    # teacher_id isn't on the answers table — keep the column shape so the
    # filter UI still works when we wire it up.
    if "teacher_id" not in df.columns:
        df["teacher_id"] = None
    return df


@st.cache_data(ttl=300)
def load_camera_lookup() -> dict[str, dict]:
    """{camera_id: {centre_name, subject}} from cctv_cameras.xlsx."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(CAMERAS_XLSX, data_only=True)
    except Exception:
        return {}
    ws = wb["cameras"] if "cameras" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        cam = rec.get("camera_id")
        if not cam:
            continue
        out[str(cam)] = {
            "centre_name": (rec.get("centre_name") or "").strip() or "—",
        }
    return out


@st.cache_data(ttl=60)
def canonical_session_runs(answers_df: pd.DataFrame) -> pd.DataFrame:
    """Pick the best (subject, session_id) → one canonical run.

    Rule: Shape B only, latest run_n per session. activity_context is no
    longer a tier-breaker (in practice every Shape B run carries one and
    we don't have that flag in Supabase yet — kept simple).
    """
    if answers_df.empty:
        return answers_df

    cols = ["subject", "session_id", "session_date", "camera", "teacher_id",
            "rubric_version", "reasoner", "shape", "run_id", "run_n"]
    available = [c for c in cols if c in answers_df.columns]
    runs = answers_df[available].drop_duplicates().copy()
    if "shape" in runs.columns:
        runs = runs[runs["shape"] == "B"]
    if runs.empty:
        return runs
    if "run_n" in runs.columns:
        runs = runs.sort_values(
            ["subject", "session_id", "run_n"], ascending=[True, True, False],
        )
    return runs.drop_duplicates(subset=["subject", "session_id"], keep="first")


def trimmed_video_path(subject: str, session_id: str) -> Path | None:
    p = SESSIONS_DIR / subject / session_id / "3_trimmed.mp4"
    return p if p.exists() else None


def ffprobe_duration_seconds(path: Path) -> float | None:
    """ffprobe the file duration in seconds. Cached to avoid re-probing."""
    import subprocess
    try:
        r = subprocess.run(
            ["ffprobe", "-v", "error", "-show_format", "-print_format", "json", str(path)],
            capture_output=True, text=True, timeout=15,
        )
        return float(json.loads(r.stdout)["format"]["duration"])
    except Exception:
        return None


@st.cache_data(ttl=300)
def cached_video_duration(path_str: str) -> float | None:
    return ffprobe_duration_seconds(Path(path_str))


def hms_to_seconds(ts: str) -> int:
    try:
        parts = [int(x) for x in str(ts).split(":")]
        if len(parts) == 3:
            return parts[0] * 3600 + parts[1] * 60 + parts[2]
        if len(parts) == 2:
            return parts[0] * 60 + parts[1]
    except Exception:
        pass
    return 0


def parse_class_window(
    subject: str, session_id: str, trimmed_path: Path | None,
    *, run_id: str | None = None,
) -> tuple[str, str, str] | None:
    """Return (start_HHMM, end_HHMM, source_note) for the class window.

    Start: HHMM token from session_id (scheduled start).
    End  : computed from trim duration + boundary-detected flag.

    Resolution order for duration + boundary flag:
      1. The Runs sheet of tqm_answers.xlsx (works in the cloud build
         where the video isn't on disk). Pass `run_id` to enable.
      2. The video file on disk + 2_boundaries.json sibling. Fallback for
         older Runs rows that don't have the new columns.
      3. None (returns "—" for end with an explanatory note).
    """
    try:
        parts = session_id.split("__")
        hhmm = parts[-1]
        h, m = int(hhmm[:2]), int(hhmm[2:])
        start = time(h, m)
    except Exception:
        return None

    # Try the xlsx Runs sheet first
    dur_sec: float | None = None
    bd: bool | None = None
    if run_id:
        info = runs_window_info().get(str(run_id))
        if info:
            dur_sec = info.get("duration_sec")
            bd = info.get("boundaries_detected")

    # Fallback: disk inspection (local dev)
    if dur_sec is None and trimmed_path is not None and trimmed_path.exists():
        dur_sec = cached_video_duration(str(trimmed_path))
        if bd is None:
            bd = (trimmed_path.parent / "2_boundaries.json").exists()

    if dur_sec is None:
        return (start.strftime("%H:%M"), "—", "trim metadata unavailable")

    dur_min = max(1, int(round(dur_sec / 60)))
    if bd:
        class_min = dur_min
        source = "boundary-detected"
    else:
        class_min = max(1, dur_min - 2 * TRIM_MARGIN_MIN)
        source = f"manual trim (−{2*TRIM_MARGIN_MIN}min margin)"
    end_dt = datetime.combine(date.today(), start) + timedelta(minutes=class_min)
    return (start.strftime("%H:%M"), end_dt.time().strftime("%H:%M"), source)


# ─── App ───────────────────────────────────────────────────────────────────


def main():
    if not password_gate():
        return
    st.title("Openhouse")

    # ── Top-level view switcher ──
    view = st.sidebar.radio(
        "View",
        ["Sessions", "Coaching Queue"],
        index=0,
        key="_view",
    )
    st.sidebar.markdown("---")

    if view == "Coaching Queue":
        render_coaching_queue()
        return

    df = load_all_answers()
    if df.empty:
        st.warning(
            f"No data found in `{ANSWERS_XLSX.relative_to(ROOT)}`. "
            "Run `scripts/run_rubric.py` on at least one session first."
        )
        return

    canon = canonical_session_runs(df)
    cam_lookup = load_camera_lookup()

    canon = canon.copy()
    canon["centre"] = canon["camera"].map(
        lambda c: cam_lookup.get(str(c), {}).get("centre_name", "—")
    )
    canon["teacher_display"] = canon["teacher_id"].apply(
        lambda t: t if pd.notna(t) and str(t) not in ("None", "nan", "") else "—"
    )

    # ── Sidebar: filters + sessions list ──
    # All filters are OPTIONAL. Empty multiselect = "no constraint" (don't
    # apply that filter). Date range left blank = full available range.
    st.sidebar.header("Filters")

    all_subjects = sorted(canon["subject"].unique())
    sel_subjects = st.sidebar.multiselect(
        "Subject", all_subjects, default=[], placeholder="Any",
    )

    valid_dates = sorted(canon["session_date"].dropna().unique())
    if valid_dates:
        date_range = st.sidebar.date_input(
            "Date range",
            value=(),
            min_value=valid_dates[0],
            max_value=valid_dates[-1],
        )
        if isinstance(date_range, tuple) and len(date_range) == 2:
            d_lo, d_hi = date_range
        else:
            d_lo, d_hi = None, None
    else:
        d_lo, d_hi = None, None

    all_centres = sorted(canon["centre"].dropna().unique())
    sel_centres = st.sidebar.multiselect(
        "Centre", all_centres, default=[], placeholder="Any",
    )

    all_teachers = sorted(canon["teacher_display"].dropna().unique())
    sel_teachers = st.sidebar.multiselect(
        "Teacher", all_teachers, default=[], placeholder="Any",
    )

    # Apply each filter only when the user actually set it
    filtered = canon
    if sel_subjects:
        filtered = filtered[filtered["subject"].isin(sel_subjects)]
    if sel_centres:
        filtered = filtered[filtered["centre"].isin(sel_centres)]
    if sel_teachers:
        filtered = filtered[filtered["teacher_display"].isin(sel_teachers)]
    if d_lo is not None and d_hi is not None:
        filtered = filtered[
            (filtered["session_date"] >= d_lo)
            & (filtered["session_date"] <= d_hi)
        ]
    filtered = filtered.sort_values("session_date", ascending=False)

    st.sidebar.markdown("---")
    st.sidebar.metric("Sessions", len(filtered))

    if filtered.empty:
        st.sidebar.info("No sessions match the current filters.")
        st.info("Use the filters in the sidebar to load a session.")
        return

    st.sidebar.subheader("Pick a session")
    labels: list[str] = []
    keys: list[tuple[str, str, str]] = []
    for _, row in filtered.iterrows():
        sd = row["session_date"]
        sd_str = sd.strftime("%Y-%m-%d") if isinstance(sd, date) else str(sd)
        label = (
            f"**{row['subject']}** · {sd_str}\n\n"
            f"{row['centre']} · {row['teacher_display']}"
        )
        labels.append(label)
        keys.append((row["subject"], row["session_id"], row["run_id"]))

    # Default selection — if the user clicked "Open session" in the queue,
    # land on that one.
    default_idx = 0
    jump_target = st.session_state.pop("_jump_session", None)
    if jump_target is not None:
        try:
            default_idx = keys.index(jump_target)
        except ValueError:
            default_idx = 0

    idx = st.sidebar.radio(
        "Pick a session",
        range(len(labels)),
        index=default_idx,
        format_func=lambda i: labels[i],
        label_visibility="collapsed",
    )
    sel_subject, sel_sid, sel_run_id = keys[idx]

    # ── Main pane: full-width session detail ──
    render_session_detail(df, sel_subject, sel_sid, sel_run_id, cam_lookup)


def render_session_detail(
    df: pd.DataFrame, subject: str, session_id: str, run_id: str,
    cam_lookup: dict[str, dict],
):
    qdf = df[(df["subject"] == subject) & (df["run_id"] == run_id)].copy()
    # materials live on runs.materials_seen in Supabase.
    materials = materials_by_run().get(str(run_id)) or []

    cam = qdf["camera"].iloc[0] if not qdf.empty else ""
    centre = cam_lookup.get(str(cam), {}).get("centre_name", "—")
    teacher = qdf["teacher_id"].iloc[0] if not qdf.empty else None
    teacher_display = teacher if pd.notna(teacher) and str(teacher) not in ("None", "nan", "") else "—"
    sess_date = qdf["session_date"].iloc[0] if not qdf.empty else None
    sess_date_str = sess_date.strftime("%Y-%m-%d") if isinstance(sess_date, date) else str(sess_date)

    # Header (no session_id visible)
    st.subheader(f"{subject} · {sess_date_str}")
    st.caption(f"Centre: **{centre}**  ·  Teacher: **{teacher_display}**")

    # ── Anchor for scroll-to-video, then persistent video player ──
    st.markdown('<div id="video-anchor"></div>', unsafe_allow_html=True)
    v_path = trimmed_video_path(subject, session_id)
    video_state_key = f"video_jump_{session_id}_{run_id}"
    jump_to = int(st.session_state.get(video_state_key, 0))

    # Source order:
    #   1. R2 signed URL (when R2 secrets are configured — works on cloud)
    #   2. Local file (works on the Mac during dev)
    #   3. No source — show a contextual info message
    video_src = r2_signed_video_url(subject, session_id)
    if video_src is None and v_path is not None:
        video_src = str(v_path)

    if video_src is not None:
        st.video(video_src, start_time=jump_to)
        # ALWAYS render the caption (placeholder when no jump) so the
        # widget tree stays positionally stable across reruns — otherwise
        # st.expander state below resets when a timestamp is clicked.
        if jump_to:
            ts_str = f"{jump_to // 3600:02d}:{(jump_to // 60) % 60:02d}:{jump_to % 60:02d}"
            st.caption(f"Player seeking to `{ts_str}`. Tap another timestamp to jump again.")
        else:
            st.caption(" ")
    elif is_cloud_deploy():
        st.info(
            "🎬 **Video for this session isn't on R2 yet.** "
            "Add `<subject>/<session_id>.mp4` to the bucket — or open the "
            "local copy on the team Mac to play it. The Q&A, materials, and "
            "class window below are the same in both."
        )
    else:
        st.info(
            "No trimmed video on disk for this session. "
            "Run the pipeline (or manually trim) to populate."
        )

    # ALWAYS render the scroll injector for the same reason — pass do_scroll
    # so it knows whether to actually scroll on this render.
    components_html_scroll_to_anchor(
        "video-anchor",
        do_scroll=bool(st.session_state.pop("_scroll_to_video", False)),
    )

    # ── Class window + materials seen, side-by-side under the video ──
    window = parse_class_window(subject, session_id, v_path, run_id=run_id)

    m_col, w_col = st.columns([3, 2], gap="large")
    with w_col:
        st.markdown("##### Class window")
        if window:
            st.markdown(f"**Start:** `{window[0]}`")
            st.markdown(f"**End:** `{window[1]}`")
            st.caption(f"_{window[2]}_")
        else:
            st.markdown("—")

    with m_col:
        st.markdown(f"##### Materials seen ({len(materials)})")
        if not materials:
            if is_cloud_deploy():
                st.caption(
                    "_Materials list lives in the per-run JSON; available in "
                    "the local copy of the dashboard._"
                )
            else:
                st.caption(
                    "_Reasoner didn't emit `materials_seen` for this run "
                    "(older prompt version)._"
                )
        else:
            items = [str(m.get("item", "")).strip() for m in materials if m.get("item")]
            st.markdown(", ".join(items) if items else "—")

    st.markdown("---")

    # ── Section tabs: Environment / Content Knowledge / Facilitation / Warmth ──
    tabs = st.tabs([f"{SECTION_ICONS.get(s, '')} {s}" for s in SECTION_TABS])
    for tab, section_name in zip(tabs, SECTION_TABS):
        with tab:
            render_section_questions(qdf, section_name, video_state_key)


def components_html_scroll_to_anchor(anchor_id: str, *, do_scroll: bool = True) -> None:
    """Always-rendered zero-height JS injector. When `do_scroll` is True it
    scrolls Streamlit's parent document up to the element with the given id;
    when False it injects a no-op (kept so the widget tree stays stable
    across reruns — st.expander positional identity is fragile).
    """
    import streamlit.components.v1 as components
    if not do_scroll:
        components.html("<!-- noop -->", height=0)
        return
    components.html(
        f"""
        <script>
          const t = window.parent.document.getElementById("{anchor_id}");
          if (t) {{
            t.scrollIntoView({{behavior: "smooth", block: "start"}});
          }} else {{
            const c = window.parent.document.querySelector('section.main, [data-testid="stAppViewContainer"]');
            if (c) c.scrollTo({{top: 0, behavior: "smooth"}});
          }}
        </script>
        """,
        height=0,
    )


def render_section_questions(qdf: pd.DataFrame, section_name: str, video_state_key: str):
    section_df = qdf[qdf["section"] == section_name].copy()
    if section_df.empty:
        st.info(f"No questions tagged '{section_name}' for this run.")
        return

    def _qnum(qid):
        try:
            return int(str(qid).lstrip("Q"))
        except ValueError:
            return 9999
    section_df["_n"] = section_df["question_id"].map(_qnum)
    section_df = section_df.sort_values("_n")

    ans = int((~section_df["insufficient_information"]).sum())
    total = len(section_df)
    st.caption(f"{ans}/{total} answered · {total - ans} INSUFFICIENT")

    for _, row in section_df.iterrows():
        qid = row["question_id"]
        q_text = row.get("question_text") or ""
        ans_val = row.get("answer") or ""
        conf = str(row.get("confidence") or "").lower()
        ins = bool(row.get("insufficient_information"))
        ts_raw = row.get("evidence_timestamps") or ""
        ts_list = [t.strip() for t in str(ts_raw).split(",") if t.strip()] if ts_raw else []
        rationale = row.get("rationale") or ""

        conf_badge = {"high": "🟢", "medium": "🟡", "low": "🔴"}.get(conf, "⚪")
        chip = "❌ INSUFFICIENT" if ins else ans_val

        with st.expander(f"{qid} · {q_text}  —  {conf_badge} **{chip}**", expanded=False):
            cols = st.columns([3, 2])
            with cols[0]:
                st.markdown(f"**Answer:** {ans_val}")
                st.markdown(f"**Confidence:** {conf or '—'}")
                if rationale:
                    st.markdown(f"**Rationale:** {rationale}")
            with cols[1]:
                if ts_list:
                    st.markdown("**Evidence — tap to seek:**")
                    for ts in ts_list:
                        sec_off = hms_to_seconds(ts)
                        if st.button(
                            f"▶ {ts}",
                            key=f"jump_{video_state_key}_{qid}_{ts}",
                            help="Seek the video player above to this moment",
                        ):
                            st.session_state[video_state_key] = sec_off
                            st.session_state["_scroll_to_video"] = True
                            st.rerun()
                else:
                    st.markdown("_no evidence timestamps_")


# ─── Coaching Queue ──────────────────────────────────────────────────────────


# Phrases (case-insensitive) that mark a yes_no question as a "red flag"
# question — Yes here is concerning (vs e.g. "Was every child addressed?"
# where Yes is good). Matches the three Warmth-section safety questions
# across all subjects.
_RED_FLAG_QUESTION_PREFIX = "are there any instances where"

COACHING_STATUS_OPTIONS = [
    ("training_required_immediately", "🚨 Training required immediately"),
    ("training_required",             "⚠️ Training required"),
    ("no_training_required",          "✓ No training required"),
]
COACHING_STATUS_LABEL = dict(COACHING_STATUS_OPTIONS)


@st.cache_data(ttl=30)
def load_coaching_actions() -> dict[str, dict]:
    """{session_id → coaching_actions row}. Empty when Supabase isn't set up."""
    rows = _fetch_all("coaching_actions")
    return {r["session_id"]: r for r in rows}


def save_coaching_action(session_id: str, status: str, notes: str | None) -> bool:
    """Upsert a coaching_actions row. Returns True on success."""
    sb = _supabase()
    if sb is None:
        return False
    try:
        sb.table("coaching_actions").upsert({
            "session_id": session_id,
            "status": status,
            "notes": (notes or "").strip() or None,
            "updated_at": datetime.utcnow().isoformat(),
        }, on_conflict="session_id").execute()
        load_coaching_actions.clear()  # invalidate cache so the table refreshes
        return True
    except Exception as e:
        st.error(f"Save failed: {e}")
        return False


def _flag_session(qdf: pd.DataFrame) -> list[str]:
    """Apply the coaching ruleset to one canonical session's question rows.
    Returns a list of human-readable reasons. Empty list = not flagged."""
    reasons: list[str] = []
    if qdf.empty:
        return reasons

    for _, r in qdf.iterrows():
        qid = str(r.get("question_id") or "")
        ans_raw = str(r.get("answer") or "").strip()
        conf = str(r.get("confidence") or "").strip().lower()
        atype = str(r.get("answer_type") or "").strip().lower()
        ins = bool(r.get("insufficient_information"))
        qtext = str(r.get("question_text") or "")

        if ins:
            continue

        # Rule 1: low scored_1_4 with medium or high confidence
        if atype == "scored_1_4" and conf in ("medium", "high"):
            try:
                n = int(ans_raw)
                if n in (1, 2):
                    short = qtext.rstrip("?").strip()
                    if len(short) > 50:
                        short = short[:47] + "…"
                    reasons.append(f"{qid}: {short} = {n} ({conf[0].upper()})")
            except ValueError:
                pass

        # Rule 2: red-flag yes_no answered Yes (confidence not gated)
        if atype == "yes_no" and ans_raw.lower().startswith("yes"):
            if qtext.lower().startswith(_RED_FLAG_QUESTION_PREFIX):
                short = qtext.rstrip("?").strip()
                if len(short) > 50:
                    short = short[:47] + "…"
                reasons.append(f"{qid}: {short} = Yes")
    return reasons


@st.cache_data(ttl=60)
def compute_coaching_queue(df: pd.DataFrame, canon: pd.DataFrame) -> pd.DataFrame:
    """For each canonical session, evaluate the ruleset and return a frame
    of flagged sessions with a 'reasons' column."""
    if df.empty or canon.empty:
        return pd.DataFrame()
    queue_rows = []
    for _, sess in canon.iterrows():
        qdf = df[(df["subject"] == sess["subject"]) & (df["run_id"] == sess["run_id"])]
        reasons = _flag_session(qdf)
        if not reasons:
            continue
        queue_rows.append({
            "subject": sess["subject"],
            "session_id": sess["session_id"],
            "session_date": sess.get("session_date"),
            "camera": sess.get("camera"),
            "teacher_id": sess.get("teacher_id"),
            "run_id": sess["run_id"],
            "reasons": reasons,
            "flag_count": len(reasons),
        })
    return pd.DataFrame(queue_rows).sort_values("session_date", ascending=False)


def render_coaching_queue():
    st.header("Coaching Queue")
    st.caption(
        "Sessions flagged for training-team attention. A session is queued when "
        "any scored question is **1 or 2** with medium/high confidence, **or** "
        "a safety-themed yes/no question is answered **Yes** at any confidence."
    )

    sb = _supabase()
    if sb is None:
        st.error(
            "Supabase isn't configured. Add `SUPABASE_URL` and "
            "`SUPABASE_SERVICE_KEY` to Streamlit Cloud Secrets (or .env locally)."
        )
        return

    df = load_all_answers()
    canon = canonical_session_runs(df)
    cam_lookup = load_camera_lookup()
    queue = compute_coaching_queue(df, canon)
    coaching = load_coaching_actions()

    if queue.empty:
        st.success("No sessions match the coaching rules right now.")
        return

    queue = queue.copy()
    queue["centre"] = queue["camera"].map(
        lambda c: cam_lookup.get(str(c), {}).get("centre_name", "—")
    )
    queue["teacher_display"] = queue["teacher_id"].apply(
        lambda t: t if pd.notna(t) and str(t) not in ("None", "nan", "") else "—"
    )

    # ── Top-line filters ──
    cols = st.columns(3)
    with cols[0]:
        subjects = sorted(queue["subject"].unique())
        sel_subj = st.multiselect("Subject", subjects, default=[], placeholder="Any")
    with cols[1]:
        centres = sorted(queue["centre"].dropna().unique())
        sel_centre = st.multiselect("Centre", centres, default=[], placeholder="Any")
    with cols[2]:
        sel_status = st.multiselect(
            "Status",
            ["unreviewed"] + [v for _, v in COACHING_STATUS_OPTIONS],
            default=[],
            placeholder="Any",
        )

    view = queue
    if sel_subj:
        view = view[view["subject"].isin(sel_subj)]
    if sel_centre:
        view = view[view["centre"].isin(sel_centre)]
    if sel_status:
        def _row_status(sid: str) -> str:
            row = coaching.get(sid)
            if not row:
                return "unreviewed"
            return COACHING_STATUS_LABEL.get(row.get("status"), row.get("status") or "?")
        view = view[view["session_id"].apply(lambda s: _row_status(s) in sel_status)]

    st.markdown(f"**{len(view)}** session(s) flagged")
    st.markdown("---")

    # ── One row per session ──
    for _, row in view.iterrows():
        sid = row["session_id"]
        existing = coaching.get(sid) or {}
        cur_status = existing.get("status")
        cur_notes = existing.get("notes") or ""

        d = row["session_date"]
        d_str = d.strftime("%Y-%m-%d") if isinstance(d, date) else str(d)

        header_left = f"**{row['subject']}** · {d_str} · {row['centre']} · {row['teacher_display']}"
        status_badge = (
            COACHING_STATUS_LABEL.get(cur_status, "")
            if cur_status else "○ Unreviewed"
        )

        with st.expander(f"{header_left}  —  {status_badge}", expanded=False):
            # Reasons + open-session link
            top = st.columns([5, 1])
            with top[0]:
                st.markdown("**Flagged because:**")
                for r in row["reasons"]:
                    st.markdown(f"- {r}")
            with top[1]:
                if st.button("Open session", key=f"open_{sid}"):
                    st.session_state["_view"] = "Sessions"
                    st.session_state["_jump_session"] = (
                        row["subject"], sid, row["run_id"],
                    )
                    st.rerun()

            # Status + notes
            st.markdown("---")
            new_status = st.radio(
                "Decision",
                [k for k, _ in COACHING_STATUS_OPTIONS],
                format_func=lambda k: COACHING_STATUS_LABEL[k],
                index=(
                    [k for k, _ in COACHING_STATUS_OPTIONS].index(cur_status)
                    if cur_status in dict(COACHING_STATUS_OPTIONS) else 1
                ),
                key=f"status_{sid}",
                horizontal=True,
            )
            new_notes = st.text_area(
                "Notes",
                value=cur_notes,
                key=f"notes_{sid}",
                placeholder="Coaching plan, who's responsible, follow-up date…",
                height=80,
            )
            if st.button("Save", key=f"save_{sid}", type="primary"):
                if save_coaching_action(sid, new_status, new_notes):
                    st.success("Saved.")
                    st.rerun()


if __name__ == "__main__":
    main()
