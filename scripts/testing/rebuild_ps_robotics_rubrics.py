"""
One-off: rebuild the `Public Speaking` and `Robotics` tabs of
`prompts/rubrics.xlsx` to mirror the Art tab's shape (Bucket / Criteria /
Question / Answer type / Level 1-4) for age band 5-8.

Why this script exists:
  - The two tabs were stubs (only Criteria + Question, no answer types, no
    1-4 levels) — they all fell back to free_text.
  - The HTML preview "Quality Assessor & Feedback" (~2026-06-29) revealed
    the canonical 5-8 segment names: PS = Roll Call / Playground /
    Showtime / Experience Book; Robotics = Experiments / Builds /
    Experience Book.
  - Facilitation + Warmth level wording is reused verbatim from the Art
    tab so the three subjects' delivery scores compare cleanly.

Run once:
    .venv/bin/python scripts/testing/rebuild_ps_robotics_rubrics.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
WORKBOOK = ROOT / "prompts" / "rubrics.xlsx"

HEADERS = ("Bucket", "Criteria", "What AI needs to observe",
           "Answer type", "Level 1", "Level 2", "Level 3", "Level 4")

# Shared 1-4 level wording (copied verbatim from the Art tab so the three
# subjects' Facilitation/Warmth scores compare cleanly).
L_REACH = ("fixed rows; all routed through the educator",
           "mostly static; some grouping",
           "grouped, some free movement",
           "children reach each other & materials freely")
L_SETUP = ("nothing ready; scrambling after children arrive",
           "some materials out, some missing",
           "most ready, minor gaps",
           "everything out & reachable before children settle")
L_EXPLAIN_CORRECT = (
    "The teacher explained another activity / the explanation was incorrect",
    "Some parts of the activity were explained correctly but most parts were "
    "incorrectly explained, affecting how the activity will be conducted",
    "Most parts were correctly explained but some parts were incorrect, "
    "not affecting how the activity will be conducted overall",
    "The teacher explained the activity correctly",
)
L_EXPLAIN_CONFIDENT = (
    "The teacher was not confident while explaining the activity, fumbled and "
    "had to refer to the material again and again",
    "The teacher was mostly unconfident while explaining the activity",
    "The teacher was mostly confident while explaining the activity",
    "The teacher was completely confident while explaining the activity",
)
L_START = ("children unclear what to do at all",
           "some confusion; repeated re-explaining",
           "clear after a short start",
           "children know what & why within a minute")
L_STRUGGLE = ("ignores or shames the struggle",
              "notices but moves on",
              "helps, somewhat",
              "meets them, adjusts, keeps them in it")
L_ADAPT = ("one level for all; strugglers & fast finishers stuck",
           "rare adjustment",
           "adjusts for some",
           "smoothly makes it easier or harder without losing the point")
L_PROMPT = ("tells children exactly what to do throughout",
            "mostly directs, little space",
            "some open prompts & wait time",
            "asks open questions, gives wait time, lets children explore")
L_ATTENTION = ("attention to a favoured few",
               "some left out",
               "most included",
               "actively reaches the quiet ones too")
L_TRANSITION = ("chaos between parts; long time lost",
                "clunky; noticeable dead time",
                "mostly smooth",
                "brisk handovers, under a minute lost")
L_AUTHORITY = (
    "control through fear / sharpness / raising their volume",
    "wobbles between soft & harsh. Has to put in effort to hold the class",
    "mostly calm & firm. Able to hold the class for most of the time",
    "firm and calm; respect without fear",
)
L_TONE = ("cold or harsh",
          "flat / transactional",
          "generally warm",
          "consistently warm; children feel at ease")
L_REWARD = ("praises only correct answers",
            "rarely names effort",
            "often notices effort",
            "consistently rewards trying, not just success")
L_CURIOSITY = ("visibly disengaged from the subject",
               "going through the motions",
               "some genuine interest",
               "contagious enthusiasm for the subject")


def _row(bucket, criteria, question, atype, levels=None):
    """Build one workbook row. Bucket/Criteria can be None (carry-forward)."""
    h, i, j, k = (levels if levels else (None, None, None, None))
    return (bucket, criteria, question, atype, h, i, j, k)


def public_speaking_rows():
    """32 rubric questions for PS 5-8. Same shape as the Art tab.

    Segments (HTML 5-8): Roll Call -> Playground -> Showtime -> Experience Book.
    Roll Call is the warmup (sentence chain, voice toss, copycat, ...);
    Playground is the main game block (script flip, body talk, guess me, ...);
    Showtime is the performance/sharing block (mad ad, magic box, story spine,
    superhero sales pitch, ...).
    """
    rows = []
    # --- Environment: setup + materials -----------------------------------
    crit_setup = ("Classroom is setup in a way that makes it clear what the "
                  "agenda is and children are able to access games, lanyards, "
                  "and progress booklets easily")
    rows.append(_row("Environment", crit_setup,
                     "Were all the Public Speaking games and materials present?",
                     "yes_no"))
    rows.append(_row(None, None,
                     "Were the lanyards and progress booklets visible and accessible?",
                     "yes_no"))
    rows.append(_row(None, None,
                     "Were the children able to reach materials and each other "
                     "without going through the educator?",
                     "scored_1_4", L_REACH))
    rows.append(_row(None, None,
                     "Was the class setup before the children entered the classroom?",
                     "scored_1_4", L_SETUP))

    # --- Environment: segment durations -----------------------------------
    crit_struct = ("Was the class structured like a typical Public Speaking "
                   "session\nRoll Call --> Playground --> Showtime --> "
                   "Experience Book?")
    rows.append(_row("Environment", crit_struct,
                     "# of minutes spent on Roll Call?", "numeric"))
    rows.append(_row(None, None,
                     "# of minutes spent on Playground?", "numeric"))
    rows.append(_row(None, None,
                     "# of minutes spent on Showtime?", "numeric"))
    rows.append(_row(None, None,
                     "# of minutes spent on Experience Book?", "numeric"))

    # --- Content Knowledge: game explanation (correctness + confidence) ---
    crit_games = ("Were they well-versed with the Public Speaking games used "
                  "in the class (Roll Call warm-ups and Playground games)?")
    rows.append(_row("Content Knowledge", crit_games,
                     "Did the teacher explain the Playground game correctly?",
                     "scored_1_4", L_EXPLAIN_CORRECT))
    rows.append(_row(None, None,
                     "Did the teacher explain the Roll Call activity correctly?",
                     "scored_1_4", L_EXPLAIN_CORRECT))
    rows.append(_row(None, None,
                     "Did the teacher explain the Playground game confidently?",
                     "scored_1_4", L_EXPLAIN_CONFIDENT))
    rows.append(_row(None, None,
                     "Did the teacher explain the Roll Call activity confidently?",
                     "scored_1_4", L_EXPLAIN_CONFIDENT))

    # --- Content Knowledge: showtime setup + start-by-self ----------------
    crit_show = ("Were they able to break down the Showtime activity into "
                 "clear, manageable steps for students?")
    rows.append(_row("Content Knowledge", crit_show,
                     "What instructional formats did the teacher use to "
                     "explain the Showtime activity?",
                     "free_text"))
    rows.append(_row(None, None,
                     "Were the children able to start the Showtime activity "
                     "themselves after the explanation?",
                     "scored_1_4", L_START))

    # --- Facilitation: level-up/down --------------------------------------
    crit_level = "Can they level up/down for a child/children?"
    rows.append(_row("Facilitation", crit_level,
                     "# of children who were disengaged/struggling with "
                     "the activity", "numeric"))
    rows.append(_row(None, None,
                     "# of children who completed the activity within "
                     "5 min of end time", "numeric"))
    rows.append(_row(None, None,
                     "# of children who completed the activity at least "
                     "5 min before the end time", "numeric"))
    rows.append(_row(None, None,
                     "Does the teacher respond well to a child who is struggling?",
                     "scored_1_4", L_STRUGGLE))
    rows.append(_row(None, None,
                     "Is the teacher able to adapt for children at different levels?",
                     "scored_1_4", L_ADAPT))

    # --- Facilitation: engage all children --------------------------------
    crit_engage = ("Are they able to engage all children or attempt to engage "
                   "all children throughout the class?")
    rows.append(_row("Facilitation", crit_engage,
                     "Does the educator prompt and observe or do they instruct",
                     "scored_1_4", L_PROMPT))
    rows.append(_row(None, None,
                     "Was every child present individually addressed "
                     "(eye contact + name or direct prompt) at least once?",
                     "scored_1_4", L_ATTENTION))
    rows.append(_row(None, None,
                     "How does the teacher handle transitions between the segments?",
                     "scored_1_4", L_TRANSITION))

    # --- Facilitation: behaviour management -------------------------------
    crit_behav = ("Are they able to manage behaviour with a balance of being "
                  "polite and firm?")
    rows.append(_row("Facilitation", crit_behav,
                     "# of instances where there was a disruption in the class",
                     "numeric"))
    rows.append(_row(None, None,
                     "How did the teacher handle the disruption", "free_text"))
    rows.append(_row(None, None,
                     "Did the disruption end once the teacher intervened",
                     "yes_no"))
    rows.append(_row(None, None,
                     "Is the educator able to hold the room with calm, "
                     "consistent authority and without raising their volume "
                     "or through fear?",
                     "scored_1_4", L_AUTHORITY))

    # --- Warmth: child safety ---------------------------------------------
    crit_safe = "Do the children feel safe in the teacher's presence?"
    rows.append(_row("Warmth", crit_safe,
                     "Are there any instances where the child may feel unsafe?",
                     "yes_no"))
    rows.append(_row(None, None,
                     "Are there any instances where the teacher shouts or "
                     "punishes a child?", "yes_no"))
    rows.append(_row(None, None,
                     "Are there any instances where the teacher seems "
                     "frustrated, irritated, impatient?", "yes_no"))

    # --- Warmth: tone, reward, curiosity ----------------------------------
    crit_warm = ("Teacher interacts with children with affection, empathy, "
                 "patience, and respect")
    rows.append(_row("Warmth", crit_warm,
                     "How is the educator's tone?",
                     "scored_1_4", L_TONE))
    rows.append(_row(None, None,
                     "What does the educator encourage and reward in the classroom?",
                     "scored_1_4", L_REWARD))
    rows.append(_row(None, None,
                     "Does the educator model curiosity and genuine interest "
                     "in what they are teaching?",
                     "scored_1_4", L_CURIOSITY))
    return rows


def robotics_rows():
    """31 rubric questions for Robotics 5-8.

    Segments (HTML 5-8): Experiments -> Builds -> Experience Book.
    Experiments are the L1 lever / pulley demonstrations; Builds are the
    physical kit work (see-saw, weighing scale, crane).
    """
    rows = []
    # --- Environment: setup + materials -----------------------------------
    crit_setup = ("Classroom is setup in a way that makes it clear what the "
                  "agenda is: experiment materials + cue cards, and build kits + "
                  "reference models are all out and reachable")
    rows.append(_row("Environment", crit_setup,
                     "Were all the experiment materials present?", "yes_no"))
    rows.append(_row(None, None,
                     "Were all the cue cards for the experiment present?",
                     "yes_no"))
    rows.append(_row(None, None,
                     "Were all the build kits and reference models present?",
                     "yes_no"))
    rows.append(_row(None, None,
                     "Were the children able to reach materials and each other "
                     "without going through the educator?",
                     "scored_1_4", L_REACH))
    rows.append(_row(None, None,
                     "Was the class setup before the children entered the classroom?",
                     "scored_1_4", L_SETUP))

    # --- Environment: segment durations -----------------------------------
    crit_struct = ("Was the class structured like a typical Robotics session\n"
                   "Experiments --> Builds --> Experience Book?")
    rows.append(_row("Environment", crit_struct,
                     "# of minutes spent on Experiments?", "numeric"))
    rows.append(_row(None, None,
                     "# of minutes spent on Builds?", "numeric"))
    rows.append(_row(None, None,
                     "# of minutes spent on Experience Book?", "numeric"))

    # --- Content Knowledge: experiment + build explanation ---------------
    crit_exp = ("Were they well-versed with the experiment they conducted? "
                "(e.g. lever / pulley — any concept-driven experiment counts)")
    rows.append(_row("Content Knowledge", crit_exp,
                     "Did the teacher explain the experiment correctly?",
                     "scored_1_4", L_EXPLAIN_CORRECT))
    rows.append(_row(None, None,
                     "Did the teacher explain the experiment confidently?",
                     "scored_1_4", L_EXPLAIN_CONFIDENT))
    rows.append(_row(None, None,
                     "Was the teacher able to answer the questions children "
                     "asked regarding the experiment?",
                     "scored_1_4", L_EXPLAIN_CORRECT))

    crit_build = ("Were they well-versed with the build that was constructed? "
                  "(e.g. see-saw / weighing scale / crane / motorised build "
                  "— any kit-based concept build counts)")
    rows.append(_row("Content Knowledge", crit_build,
                     "Did the teacher explain the build correctly?",
                     "scored_1_4", L_EXPLAIN_CORRECT))
    rows.append(_row(None, None,
                     "Did the teacher explain the build confidently?",
                     "scored_1_4", L_EXPLAIN_CONFIDENT))
    rows.append(_row(None, None,
                     "Was the teacher able to answer the questions children "
                     "asked regarding the build?",
                     "scored_1_4", L_EXPLAIN_CORRECT))

    # --- Content Knowledge: build setup + start-by-self ------------------
    crit_show = ("Were they able to break down the build process into clear, "
                 "manageable steps for students?")
    rows.append(_row("Content Knowledge", crit_show,
                     "What instructional formats did the teacher use to "
                     "explain the build?",
                     "free_text"))
    rows.append(_row(None, None,
                     "Were the children able to start the build themselves "
                     "after the explanation?",
                     "scored_1_4", L_START))

    # --- Facilitation: level-up/down --------------------------------------
    crit_level = "Can they level up/down for a child/children?"
    rows.append(_row("Facilitation", crit_level,
                     "# of children who were disengaged/struggling with the activity",
                     "numeric"))
    rows.append(_row(None, None,
                     "# of children who completed the activity within 5 min of end time",
                     "numeric"))
    rows.append(_row(None, None,
                     "# of children who completed the activity at least 5 min "
                     "before the end time", "numeric"))
    rows.append(_row(None, None,
                     "Does the teacher respond well to a child who is struggling?",
                     "scored_1_4", L_STRUGGLE))
    rows.append(_row(None, None,
                     "Is the teacher able to adapt for children at different levels?",
                     "scored_1_4", L_ADAPT))

    # --- Facilitation: engage all children --------------------------------
    crit_engage = ("Are they able to engage all children or attempt to engage "
                   "all children throughout the class?")
    rows.append(_row("Facilitation", crit_engage,
                     "Does the educator prompt and observe or do they instruct",
                     "scored_1_4", L_PROMPT))
    rows.append(_row(None, None,
                     "Was every child present individually addressed "
                     "(eye contact + name or direct prompt) at least once?",
                     "scored_1_4", L_ATTENTION))
    rows.append(_row(None, None,
                     "How does the teacher handle transitions between the segments?",
                     "scored_1_4", L_TRANSITION))

    # --- Facilitation: behaviour management -------------------------------
    crit_behav = ("Are they able to manage behaviour with a balance of being "
                  "polite and firm?")
    rows.append(_row("Facilitation", crit_behav,
                     "# of instances where there was a disruption in the class",
                     "numeric"))
    rows.append(_row(None, None,
                     "How did the teacher handle the disruption", "free_text"))
    rows.append(_row(None, None,
                     "Did the disruption end once the teacher intervened", "yes_no"))
    rows.append(_row(None, None,
                     "Is the educator able to hold the room with calm, "
                     "consistent authority and without raising their volume "
                     "or through fear?",
                     "scored_1_4", L_AUTHORITY))

    # --- Warmth: child safety ---------------------------------------------
    crit_safe = "Do the children feel safe in the teacher's presence?"
    rows.append(_row("Warmth", crit_safe,
                     "Are there any instances where the child may feel unsafe?",
                     "yes_no"))
    rows.append(_row(None, None,
                     "Are there any instances where the teacher shouts or "
                     "punishes a child?", "yes_no"))
    rows.append(_row(None, None,
                     "Are there any instances where the teacher seems "
                     "frustrated, irritated, impatient?", "yes_no"))

    # --- Warmth: tone, reward, curiosity ----------------------------------
    crit_warm = ("Teacher interacts with children with affection, empathy, "
                 "patience, and respect")
    rows.append(_row("Warmth", crit_warm,
                     "How is the educator's tone?",
                     "scored_1_4", L_TONE))
    rows.append(_row(None, None,
                     "What does the educator encourage and reward in the classroom?",
                     "scored_1_4", L_REWARD))
    rows.append(_row(None, None,
                     "Does the educator model curiosity and genuine interest "
                     "in what they are teaching?",
                     "scored_1_4", L_CURIOSITY))
    return rows


def _replace_sheet(wb: openpyxl.Workbook, name: str, rows: list[tuple]) -> None:
    """Drop and recreate `name` so we don't inherit stale rows / merges."""
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    print(f"  rebuilt {name!r}: {len(rows)} rows")


def main():
    print(f"opening {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK)

    ps_rows = public_speaking_rows()
    rob_rows = robotics_rows()

    # Re-order sheets so they appear: Art -> Public Speaking -> Robotics
    art_idx = wb.sheetnames.index("Art")
    _replace_sheet(wb, "Public Speaking", ps_rows)
    _replace_sheet(wb, "Robotics", rob_rows)
    # Move new sheets right after Art so the workbook reads naturally.
    new_order = ["Art", "Public Speaking", "Robotics"]
    wb._sheets = [wb[n] for n in new_order]

    wb.save(WORKBOOK)
    print(f"saved {WORKBOOK}")


if __name__ == "__main__":
    main()
