"""
TQM consolidation — migration step 2: bucket data/raw/ by subject.

Classifies each entry in data/raw/ by parsing the filename:

    Pattern:  D<N>_<centre>_<subject>_<YYYYMMDD>_<HHMMSS>.mp4
              (where <subject> is one of the values in cctv_cameras.xlsx)
    Action:   move to data/raw/<subject>/<file>

Any entry that doesn't match (older test files, the `playground/` dir,
test_video.mp4, etc.) moves to data/_archive/raw_pre_subject_organization/
since those predate the subject-naming convention.

Every move is logged to migration_log.txt via atomic_mv_with_log so step 2
can be reversed via `python scripts/migrate.py reverse --step 2`.

Usage:
    # Preview (default — touches nothing)
    .venv/bin/python scripts/migrate_step2_bucket_raw.py

    # Actually perform the moves
    .venv/bin/python scripts/migrate_step2_bucket_raw.py --execute

    # Skip y/N confirmation in --execute mode
    .venv/bin/python scripts/migrate_step2_bucket_raw.py --execute --yes
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from scripts.migrate import (
    PROJECT_ROOT,
    atomic_mv_with_log,
    require_confirmation,
)

STEP = "2"

RAW_DIR = PROJECT_ROOT / "data" / "raw"
ARCHIVE_DIR = PROJECT_ROOT / "data" / "_archive" / "raw_pre_subject_organization"
CAMERAS_XLSX = PROJECT_ROOT / "data" / "cctv_cameras.xlsx"

# Filename pattern that carries subject:
#   D<digits>_<centre>_<subject>_<YYYYMMDD>_<HHMMSS>.mp4
# subject may contain underscores (e.g. "public_speaking"), so we anchor on
# the trailing 8-digit + 6-digit timestamp.
SUBJECT_FILENAME_RE = re.compile(
    r"^(?P<cam>D\d+)"
    r"_(?P<centre>[a-z]+)"
    r"_(?P<subject>.+)"
    r"_(?P<date>\d{8})"
    r"_(?P<time>\d{6})"
    r"\.mp4$",
    re.IGNORECASE,
)


def load_valid_subjects() -> set[str]:
    """Read cctv_cameras.xlsx and return the set of valid subject tokens."""
    if not CAMERAS_XLSX.exists():
        raise SystemExit(f"missing camera config: {CAMERAS_XLSX}")
    wb = openpyxl.load_workbook(CAMERAS_XLSX, data_only=True)
    ws = wb["cameras"]
    headers = [c.value for c in ws[1]]
    subject_col = headers.index("subject")
    subjects: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[subject_col]
        if v:
            subjects.add(str(v).strip().lower())
    if not subjects:
        raise SystemExit(f"no subjects found in {CAMERAS_XLSX}!cameras")
    return subjects


def classify(entry: Path, valid_subjects: set[str]) -> Path | None:
    """Return the destination Path for `entry`, or None to skip it entirely.

    Skipped cases:
      - hidden entries (start with `.`) — e.g. cctv_pull's `.tmp/` workspace
      - entries whose name matches a valid subject — already bucketed from a
        previous run; idempotent
    """
    name = entry.name
    if name.startswith("."):
        return None
    if name in valid_subjects:
        return None  # already a subject bucket

    m = SUBJECT_FILENAME_RE.match(name) if entry.is_file() else None
    if m:
        subject = m.group("subject").lower()
        if subject in valid_subjects:
            return RAW_DIR / subject / name
        # Filename matches the pattern but subject token unknown → archive,
        # don't silently create a new subject bucket from an unknown token.
    return ARCHIVE_DIR / name


def _short(p: Path) -> str:
    try:
        return str(p.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(p)


def build_plan(valid_subjects: set[str]) -> tuple[list[tuple[Path, Path]], dict[str, int]]:
    """Return (moves, summary_counts)."""
    if not RAW_DIR.exists():
        raise SystemExit(f"data/raw/ not found: {RAW_DIR}")
    entries = sorted(RAW_DIR.iterdir(), key=lambda p: p.name)
    moves: list[tuple[Path, Path]] = []
    counts: Counter = Counter()
    for entry in entries:
        dst = classify(entry, valid_subjects)
        if dst is None:
            counts["skipped (hidden / already-bucketed)"] += 1
            continue
        moves.append((entry, dst))
        try:
            bucket = dst.parent.relative_to(PROJECT_ROOT)
        except ValueError:
            bucket = dst.parent
        counts[str(bucket)] += 1
    return moves, dict(counts)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Actually perform the moves (default is dry-run)",
    )
    parser.add_argument(
        "--yes",
        action="store_true",
        help="Skip the y/N confirmation in --execute mode",
    )
    parser.add_argument(
        "--show-all",
        action="store_true",
        help="Print every move (default: print first 5 + last 2 per destination)",
    )
    args = parser.parse_args()

    valid_subjects = load_valid_subjects()
    print(f"Valid subjects from cctv_cameras.xlsx: {sorted(valid_subjects)}")
    print(f"Source: {_short(RAW_DIR)}")
    print(f"Archive sink for unmatched: {_short(ARCHIVE_DIR)}")
    print()

    moves, counts = build_plan(valid_subjects)

    print("Plan summary (entries → destination):")
    for dest, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f"  {n:>4}  {dest}")
    print(f"  ----")
    print(f"  {len(moves):>4}  total moves planned")
    print()

    # Print sample of each destination bucket so user can sanity-check filename
    # parsing before committing.
    by_dest: dict[str, list[tuple[Path, Path]]] = {}
    for src, dst in moves:
        by_dest.setdefault(str(dst.parent), []).append((src, dst))
    print("Samples per destination bucket:")
    for dest, items in sorted(by_dest.items()):
        try:
            display_dest = str(Path(dest).relative_to(PROJECT_ROOT))
        except ValueError:
            display_dest = dest
        print(f"\n  [{display_dest}]  ({len(items)} entries)")
        if args.show_all or len(items) <= 7:
            for src, _ in items:
                print(f"    {src.name}")
        else:
            for src, _ in items[:5]:
                print(f"    {src.name}")
            print(f"    ... {len(items) - 7} more ...")
            for src, _ in items[-2:]:
                print(f"    {src.name}")

    if not args.execute:
        print(
            "\nDRY RUN — nothing moved. Re-run with --execute to perform the moves."
        )
        return 0

    if not require_confirmation(
        [f"About to move {len(moves)} entries from data/raw/."],
        auto_yes=args.yes,
    ):
        print("Aborted by user.")
        return 1

    print()
    moved = 0
    for src, dst in moves:
        ok = atomic_mv_with_log(src, dst, step=STEP)
        if ok:
            moved += 1

    print(f"\nDone. Moved {moved} of {len(moves)} entries.")
    print(f"To reverse: python scripts/migrate.py reverse --step {STEP} --dry-run")
    return 0


if __name__ == "__main__":
    sys.exit(main())
