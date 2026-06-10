"""
End-to-end art-class rubric test on raw NVR clips.

Pipeline:
  1. Concat 2+ raw NVR clips and transcode to H.264 + AAC at 720p
  2. Run boundary detection (class start / class end) via existing pipeline
  3. Trim to [class_start − 1 min, class_end + 1 min]
  4. Build rubric prompt from the Art sheet of the supplied Excel and run on
     the trimmed video
  5. Write per-question results to results.csv + console table

Each stage is idempotent — output cached on disk; skip on re-run unless
--force-from <stage> is passed.

Usage:
  python scripts/run_art_rubric_test.py \\
    --rubric "/Users/oh/Downloads/Teacher Quality Monitoring.xlsx" \\
    --sheet Art \\
    --segments data/raw/D28_hrbr_art_20260518_083132.mp4 \\
               data/raw/D28_hrbr_art_20260518_094903.mp4 \\
    --model gemini-3.1-pro-preview
"""

import argparse
import csv
import json
import logging
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime, date
from pathlib import Path
from typing import Optional

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl

from adapters.llm import LLMAdapter, parse_json_lenient, prompt_hash
from pipeline.boundaries import _parse_hms, _derive_elapsed_from_walls
from pipeline.render import _jinja_env, load_prompt, split_system_user
from pipeline.types import BoundaryDetection, SessionMeta

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("art_rubric_test")
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("google_genai").setLevel(logging.WARNING)


MARGIN_SECONDS = 60  # 1-min margin before class start and after class end

# Gemini does not honour the source video's encoded fps — its default sampling
# rate (~1 fps internally) blows past the 1M-token input limit on long videos.
# These caps tell Gemini to sample less aggressively for our two video calls.
BOUNDARY_FPS = 0.3      # ~1 frame every 3.3 sec, fine for "child present?" + clock reads
RUBRIC_FPS = 0.5        # ~1 frame every 2 sec, fine for events lasting multi-sec

# Filename anchor: NVR clips land as D{NN}_{centre}_{subject}_{YYYYMMDD}_{HHMMSS}.mp4.
# The trailing HHMMSS is the wall clock when the recording started — ground truth
# we can use to anchor boundary detection when Gemini misreads the clock at
# frame 0 (it sometimes returns "00:00:00" if the clock is illegible).
SEGMENT_NAME_RE = re.compile(
    r".*_(\d{8})_(\d{2})(\d{2})(\d{2})\.mp4$"
)


def derive_wall_clock_from_filename(path: Path) -> Optional[str]:
    """Return 'HH:MM:SS' parsed from a NVR-style filename, or None."""
    m = SEGMENT_NAME_RE.match(path.name)
    if not m:
        return None
    _, hh, mm, ss = m.groups()
    return f"{hh}:{mm}:{ss}"


# ─── Rubric parsing ─────────────────────────────────────────────────────────

def parse_rubric_sheet(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Parse the rubric Excel sheet into a flat list of questions.

    The sheet has 'merged-like' layout: the criteria and group description
    columns are only filled on the first row of each group; subsequent rows
    in the same group leave them blank. We fill them down so every question
    carries its full context.

    Expected columns (after the auto-numbered first column):
      criteria | group_description | question | input_required | analysis
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    questions = []
    cur_criteria = None
    cur_group = None
    qid = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        # Some sheets have leading blank column; detect dynamically
        # We expect: col A may be blank, col B = criteria, col C = group_desc,
        # col D = question, col E = input_req, col F = analysis  (Art sheet has 5 data cols)
        # Looking at the Art sheet structure, col 0 is blank or carries criteria.
        # We treat col 0 as criteria, col 1 as group, col 2 as question, col 3 as input, col 4 as analysis.
        if not row or len(row) < 3:
            continue
        criteria = row[0] if len(row) > 0 else None
        group = row[1] if len(row) > 1 else None
        question = row[2] if len(row) > 2 else None
        input_req = row[3] if len(row) > 3 else None
        analysis = row[4] if len(row) > 4 else None

        # Fill-down: empty cells inherit from the previous row
        if criteria is not None and str(criteria).strip():
            cur_criteria = str(criteria).strip()
        if group is not None and str(group).strip():
            cur_group = " ".join(str(group).split())  # collapse multi-line groups
        if question is None or not str(question).strip():
            continue  # blank row

        qid += 1
        questions.append({
            "id": f"Q{qid}",
            "criteria": cur_criteria or "(uncategorised)",
            "group": cur_group or "",
            "question": " ".join(str(question).split()),
            "input_required": " ".join(str(input_req).split()) if input_req else "",
            "analysis": " ".join(str(analysis).split()) if analysis else "",
        })
    return questions


# ─── Stage 1: concat + transcode ────────────────────────────────────────────

def stage1_combine(segments: list[Path], output: Path, boundary_input: Path,
                   force: bool = False) -> None:
    """Build two outputs:
    1a. `output` (combined.mp4): full quality, 720p, all frames — used for the
        final rubric pass (after trimming).
    1b. `boundary_input` (boundary_input.mp4): 0.5 fps, 480p, no audio — used
        ONLY for boundary detection. Long classroom recordings at full fps blow
        through Gemini's 1M-token input limit; the wall clock + child presence
        are still legible at this fidelity, so the boundary call fits.
    """
    if output.exists() and boundary_input.exists() and not force:
        log.info(f"[1] combine: both outputs exist, skipping")
        return
    if any(not s.exists() for s in segments):
        missing = [s for s in segments if not s.exists()]
        raise FileNotFoundError(f"Missing source segments: {missing}")

    output.parent.mkdir(parents=True, exist_ok=True)

    # Use the concat filter so we can re-encode (raw clips may be HEVC + µ-law).
    # Scale to 720p, CRF 28, AAC 96k. Keeps the combined file under Gemini's
    # ~2 GB Files API limit for 2.5+ hour windows.
    #
    # Important: when filter_complex produces labelled outputs we cannot ALSO
    # use `-vf`. The scale must be chained inside filter_complex.
    n = len(segments)
    inputs_args: list[str] = []
    for s in segments:
        inputs_args += ["-i", str(s)]
    concat_streams = "".join(f"[{i}:v:0][{i}:a:0]" for i in range(n))
    filter_complex = (
        f"{concat_streams}concat=n={n}:v=1:a=1[cv][a];"
        f"[cv]scale=1280:720[v]"
    )

    # ── 1a: high-quality combined (for trim + rubric) ──
    if not output.exists() or force:
        log.info(f"[1a] combine: {len(segments)} clip(s) → {output.name} (5–15 min)")
        log_a = output.parent / "1a_ffmpeg_combine.log"
        cmd_a = [
            "ffmpeg", "-y",
            *inputs_args,
            "-filter_complex", filter_complex,
            "-map", "[v]", "-map", "[a]",
            "-c:v", "libx264", "-crf", "28", "-preset", "fast",
            "-c:a", "aac", "-b:a", "96k",
            "-movflags", "+faststart",
            str(output),
        ]
        t0 = time.time()
        with open(log_a, "w") as logf:
            res = subprocess.run(cmd_a, stdout=logf, stderr=subprocess.STDOUT)
        if res.returncode != 0:
            tail = log_a.read_text().splitlines()[-40:]
            log.error(f"[1a] ffmpeg failed (exit {res.returncode}). Tail of {log_a.name}:")
            for line in tail:
                log.error(f"  {line}")
            raise RuntimeError(f"ffmpeg combine failed; see {log_a}")
        sz_mb = output.stat().st_size / 1e6
        log.info(f"[1a] done in {(time.time()-t0)/60:.1f} min, {sz_mb:.0f} MB")
    else:
        log.info(f"[1a] combine: {output.name} exists, skipping")

    # ── 1b: low-fps low-res silent copy for boundary detection ──
    # 0.5 fps + 480p + no audio keeps the file under Gemini's ~1M input-token
    # ceiling for 2.5-hour windows. Wall clock + child silhouettes are still
    # clearly visible at this fidelity.
    if not boundary_input.exists() or force:
        log.info(f"[1b] boundary input: re-encoding to 0.5 fps / 480p / silent → {boundary_input.name}")
        log_b = output.parent / "1b_ffmpeg_boundary_input.log"
        cmd_b = [
            "ffmpeg", "-y",
            "-i", str(output),
            "-vf", "fps=0.5,scale=854:480",
            "-an",  # strip audio (boundary detection doesn't need it)
            "-c:v", "libx264", "-crf", "30", "-preset", "fast",
            "-movflags", "+faststart",
            str(boundary_input),
        ]
        t0 = time.time()
        with open(log_b, "w") as logf:
            res = subprocess.run(cmd_b, stdout=logf, stderr=subprocess.STDOUT)
        if res.returncode != 0:
            tail = log_b.read_text().splitlines()[-40:]
            log.error(f"[1b] ffmpeg failed (exit {res.returncode}). Tail of {log_b.name}:")
            for line in tail:
                log.error(f"  {line}")
            raise RuntimeError(f"ffmpeg boundary-input failed; see {log_b}")
        sz_mb = boundary_input.stat().st_size / 1e6
        log.info(f"[1b] done in {(time.time()-t0)/60:.1f} min, {sz_mb:.0f} MB")
    else:
        log.info(f"[1b] boundary input: {boundary_input.name} exists, skipping")


def video_duration_seconds(path: Path) -> float:
    out = subprocess.check_output(
        ["ffprobe", "-v", "error",
         "-show_entries", "format=duration",
         "-of", "default=noprint_wrappers=1:nokey=1", str(path)],
    ).decode().strip()
    return float(out)


# ─── Stage 2: boundary detection ────────────────────────────────────────────

def stage2_boundaries(boundary_input: Path, full_combined: Path, run_dir: Path,
                      llm: LLMAdapter, filename_anchor: Optional[str] = None,
                      force: bool = False) -> dict:
    """Run boundary detection on the boundary_input file. Inlined here (rather
    than reusing pipeline.detect_boundaries) so we can pass fps=BOUNDARY_FPS
    through to call_gemini_video — long videos otherwise blow Gemini's 1M
    input-token limit even with low-fps source encoding.

    Duration metadata uses the FULL combined.mp4 span so the prompt's
    time-budget reasoning stays correct (boundary_input is just a sampled
    copy — same wall-clock span)."""
    out = run_dir / "2_boundaries.json"
    if out.exists() and not force:
        log.info(f"[2] boundaries: {out.name} exists, skipping")
        return json.loads(out.read_text())

    dur_sec = video_duration_seconds(full_combined)
    dur_min = int(round(dur_sec / 60))
    log.info(f"[2] boundaries: detecting on {boundary_input.name} "
             f"({dur_min} min, fps={BOUNDARY_FPS}) — Gemini call")

    session = SessionMeta(
        session_id=f"art_rubric_{run_dir.name}",
        recorded_at=date.today(),
        duration_minutes=dur_min,
        subject="art class (preschool)",
        video_path=boundary_input,
    )

    # Render the existing detect_boundaries prompt
    template_text = load_prompt("detect_boundaries")
    rendered = _jinja_env().from_string(template_text).render(
        session=session.model_dump(mode="json"),
    )
    system, user = split_system_user(rendered)
    full_prompt = f"{system}\n\n{user}"

    # Upload + call Gemini with fps capping
    video_file = llm.upload_video(boundary_input)
    raw = llm.call_gemini_video(
        prompt=full_prompt,
        video_file=video_file,
        fps=BOUNDARY_FPS,
    )
    (run_dir / "2_boundaries_raw.txt").write_text(raw)

    parsed = parse_json_lenient(raw)
    result = BoundaryDetection.model_validate(parsed)
    result.session_id = session.session_id
    result.source_model = llm.vision_model
    result.prompt_hash = prompt_hash(template_text)

    # If Gemini failed to read the clock at frame 0 (returns null or
    # "00:00:00") but we have a filename anchor (e.g. the NVR filename
    # `D28_..._083132.mp4` literally encodes wall clock 08:31:32), use the
    # filename as ground truth. Sidesteps the systemic issue where the
    # boundary_input's downsampled resolution makes the burned-in clock
    # illegible to the model at the very first frame.
    if filename_anchor and (
        not result.video_start_wall_clock
        or result.video_start_wall_clock in ("00:00:00", "0:00:00")
    ):
        log.warning(
            f"[2] Gemini returned video_start_wall_clock="
            f"{result.video_start_wall_clock!r} — overriding with filename anchor "
            f"{filename_anchor!r}"
        )
        result.video_start_wall_clock = filename_anchor

    # Derive elapsed times from wall-clock readings if present (same logic
    # as pipeline.detect_boundaries — the model reads, the code maths)
    _derive_elapsed_from_walls(result, session.session_id)

    payload = result.model_dump(mode="json")
    out.write_text(json.dumps(payload, indent=2))
    log.info(f"[2] boundaries: first_child={payload.get('first_child_visible_at')} "
             f"last_child={payload.get('last_child_visible_at')} "
             f"confidence={payload.get('confidence')}")
    if payload.get('first_child_evidence'):
        log.info(f"[2] first_child evidence: {payload['first_child_evidence']}")
    if payload.get('last_child_evidence'):
        log.info(f"[2] last_child evidence: {payload['last_child_evidence']}")
    return payload


def compute_trim_window(boundaries: dict, combined_dur_sec: float) -> tuple[float, float]:
    """Return (start_sec, end_sec) to trim, with MARGIN_SECONDS padding.
    Falls back to a safe window if boundaries are missing."""
    first = boundaries.get("first_child_visible_at")
    last = boundaries.get("last_child_visible_at")

    first_sec = _parse_hms(first) if first else None
    last_sec = _parse_hms(last) if last else None

    if first_sec is None and last_sec is None:
        log.warning("[2] no boundary info — trimming whole video (no margins)")
        return 0.0, combined_dur_sec

    start = max(0.0, (first_sec - MARGIN_SECONDS) if first_sec is not None else 0.0)
    end = min(combined_dur_sec, (last_sec + MARGIN_SECONDS) if last_sec is not None else combined_dur_sec)

    if end <= start:
        log.warning(f"[2] computed window degenerate (start={start}, end={end}) — using whole video")
        return 0.0, combined_dur_sec
    return start, end


# ─── Stage 3: trim ──────────────────────────────────────────────────────────

def stage3_trim(combined: Path, start: float, end: float, output: Path, force: bool = False) -> None:
    if output.exists() and not force:
        log.info(f"[3] trim: {output.name} exists, skipping")
        return
    log.info(f"[3] trim: {start:.0f}s → {end:.0f}s ({(end-start)/60:.1f} min) → {output.name}")
    log_path = output.parent / "3_ffmpeg_trim.log"
    # Re-encode at 1 fps + 720p to keep the rubric pass under Gemini's 1M
    # input-token limit even for ~90-min trimmed windows. Audio kept at full
    # sample rate because many rubric questions are [Audio] / [Audio - tone].
    cmd = [
        "ffmpeg", "-y",
        "-ss", f"{start}", "-to", f"{end}",
        "-i", str(combined),
        "-vf", "fps=1,scale=1280:720",
        "-c:v", "libx264", "-crf", "23", "-preset", "fast",
        "-c:a", "aac", "-b:a", "96k",
        "-movflags", "+faststart",
        str(output),
    ]
    t0 = time.time()
    with open(log_path, "w") as logf:
        result = subprocess.run(cmd, stdout=logf, stderr=subprocess.STDOUT)
    if result.returncode != 0:
        tail = log_path.read_text().splitlines()[-40:]
        log.error(f"[3] ffmpeg failed (exit {result.returncode}). Tail of {log_path.name}:")
        for line in tail:
            log.error(f"  {line}")
        raise RuntimeError(f"ffmpeg trim failed; see {log_path}")
    sz_mb = output.stat().st_size / 1e6
    log.info(f"[3] trim: done in {(time.time()-t0)/60:.1f} min, {sz_mb:.0f} MB")


# ─── Stage 4: rubric pass ───────────────────────────────────────────────────

RUBRIC_PROMPT_TEMPLATE = """You are analysing a recording of an Openhouse Pre-School Art class for preschoolers (ages 3–5).

═════════════════════════════════════════════════════════════════
VIDEO METADATA (READ FIRST — your phase enumeration MUST respect these)
═════════════════════════════════════════════════════════════════

  Video duration:    **{duration_str}** ({duration_sec} seconds total)
  First timestamp:   `00:00:00`   (CCTV wall clock: **{wallclock_start} IST**)
  Last timestamp:    `{duration_str}`   (CCTV wall clock: **{wallclock_end} IST**)

The CCTV wall clock is burned into every frame as an on-screen overlay.
Read it directly. Every timestamp in your output uses the format
`WALLCLOCK (VIDEO_OFFSET)` — see the OUTPUT FORMAT section below for
the exact rule.

The video runs the full duration above. Children appear throughout and
do activities for the bulk of this time. Your `phases` array's LAST
phase must end at (or within 30 seconds of) `{wallclock_end} ({duration_str})`
— not sooner. The single most common failure mode is to enumerate
phases for the first ~25 % of the video and stop; do not do this.


═════════════════════════════════════════════════════════════════
GLOSSARY — Openhouse Art-class structure
═════════════════════════════════════════════════════════════════

A full Art session runs for 90 minutes across five segments, in this
order. Some segments may be skipped or rearranged; identify what
actually occurred.

The programme builds three skills across all sessions: Fine Motor
(Tool Precision, Tracing, Drawing Figures & Patterns), Colour
(Explores freely, Names and recognises, Mixes and notices), and
Creative Expression (Explores Artistic Concepts, Emotional
Expression, Integrating Artistic Choices).

  ART GYM (15 min)
    A daily warm-up that builds Fine Motor and Creative Expression
    through short, focused mark-making. Rotates between two resources
    on consecutive sessions.
      - Art Gym Book — one page per session showing a pattern, a path,
        or a mark sequence for the child to complete or extend.
        Materials (child's choice): erasable markers, Play-Doh, thread,
        sequins.
        Skills built: Fine Motor — Tracing (primary), Drawing Figures
        and Patterns (secondary); Creative Expression — Integrating
        Artistic Choices (the small intentional decision of which
        material to choose).
      - Scribble Book — A4 spiral-bound book; each page shows a
        partially-complete scene with a single prompt at the bottom.
        Child draws their response. There is no correct response.
        Materials: erasable markers, Play-Doh, thread, sequins.
        Skills built: Fine Motor — Drawing Figures and Patterns
        (primary), Tool Precision (secondary, when sequins/thread are
        chosen); Creative Expression — Explores Artistic Concepts,
        Integrating Artistic Choices.
    Teacher does not teach or correct during Art Gym — they circulate
    and name what they see.

  ART GAMES (25 min — ONE GAME PER SESSION)
    Develops one art skill through purposeful, structured play. Rules
    explained once the first time a game is played, then the teacher
    sets up, steps back, and observes — no teaching during the game.
    Games rotate so Fine Motor, Colour, and Creative Expression games
    each appear regularly.
      Fine Motor games:
        - Shape Stitch — children sew through stitching templates with
          shoelaces (running, whip, or chain stitch by difficulty).
          Materials: stitching templates, shoelaces.
          Builds: Tool Precision and Tracing.
        - Stitch Me — children thread beads in a specific sequence
          (by colour, colour-and-number, or riddle).
          Materials: beads, sequence prompt cards.
          Builds: Tool Precision.
        - Magna Tiles — children build structures with Magna-Tiles
          guided by prompt cards (5 difficulty levels, 2D to 3D).
          Materials: Magna-Tiles, prompt cards.
          Builds: Tool Precision and Drawing Figures; also touches
          Integrating Artistic Choices.
        - Shape Mats — Tile Placement — children place matching shape
          tiles on large illustrated mats, or build arrangements freely.
          Materials: shape mats, shape tiles.
          Builds: Tool Precision and Tracing.
      Colour games:
        - Match Me — children match coloured beads/objects to grid
          squares using pattern cards (3 variations: simultaneous
          Colour Match, Pattern Memory, Speed Match).
          Materials: coloured beads or objects, grid, pattern cards.
          Builds: Explores freely and Names and recognises.
        - Mix It Up — children sort 30 object cards by colour using 6
          colour swatch cards as category markers.
          Materials: 30 object cards, 6 colour swatch cards.
          Builds: Names and recognises and Mixes and notices.
        - The Game of Red, Yellow and Blue — shape tiles in primary
          and secondary colours; children build tile equations
          (e.g. 2 red + 1 yellow) to discover secondary colours
          (3 variations: Build the Mix, Story Mix, Predict and Build).
          Materials: primary and secondary colour shape tiles.
          Builds: Mixes and notices and Colour Integration ★.
      Creative Expression games:
        - MiniArtventure — board game; children roll, move, land on a
          Draw / Colour / Mould / Build zone and complete a 2-minute
          art challenge from that zone's card deck (individual and
          cooperative variations).
          Materials: board game, zone cards.
          Builds: Integrating Artistic Choices and Explores Artistic
          Concepts; also touches Fine Motor and Colour.

  ARTIVERSE or ARTISTOTLE (35 min)
    The main make-something segment. Alternates between two modes;
    the two never share a session. Within each mode, activities are
    linear in order of difficulty. Builds all three skills (Fine
    Motor, Colour, Creative Expression).
      ARTIVERSE Days — children rotate through three media families;
      each project takes two sessions to complete:
        - Colourful Papers: Accordions · Circles · Mosaics · Loops
          and Chains.
          Materials: coloured paper, glue, scissors.
        - Crayons: Doodling · Colouring · Colour Mixing.
          Materials: crayons.
        - Watercolour: Hand Painting (greeting cards) · Finger
          Painting (greeting cards) · Sponge Painting · Q-tip Painting
          · Blow and Splatter.
          Materials: watercolours, brushes, sponges, Q-tips, straws.
      ARTISTOTLE Days — six illustrator-led projects, three sessions
      each. Children encounter the work of a famous illustrator
      (characters, palette, style) then make their own piece in that
      spirit. Materials vary per illustrator.
      Note: Emotional Expression through Art and Visual Arts
      Integration ★ are assessable on Artistotle Day 3 and Artiverse
      Chapter 3.

  EXPERIENCE BOOK (10 min)
    The teacher records the session in each child's personal
    Experience Book — what happened and what the child learnt — and
    the child adds one drawing of their own.
    Materials: child's personal Experience Book + their choice of
    drawing tool.
    Closes active making with quiet reflection.

  ART CARE (5 min)
    Children sort all materials back to the correct shelf sections
    and clean the making space. The standard is care, not speed.
    If a brush hasn't been cleaned correctly, the teacher shows how
    — once — without framing it as a correction.
    Materials: the materials used in the session + cleaning supplies.
    Skill built: care of tools and shared space.

═════════════════════════════════════════════════════════════════
CRITICAL: HOW TO DEFINE PHASE BOUNDARIES (READ TWICE)
═════════════════════════════════════════════════════════════════

The COMMON MISTAKE is to end a phase when the teacher STOPS EXPLAINING.
This is wrong.

A phase's `end` is when the **activity itself stops** — i.e. children
either:
  (a) pack away materials and move to a different activity, OR
  (b) the teacher signals the activity is over and moves on.

If the teacher explains the Art Game at `09:18:42 (00:05:25)` and the
children play that game until `09:43:42 (00:30:25)` (with the teacher
just walking around supporting them), the `art_games` phase is
**`09:18:42 (00:05:25)` → `09:43:42 (00:30:25)`** — NOT
`09:18:42 (00:05:25)` → `09:21:42 (00:08:25)` (when the explanation
ended). The 25 minutes of children actually playing the game ARE part
of the phase.

The same applies to Art Gym, Artiverse/Artistotle, etc. The ACTIVITY
phase includes the entire time children are doing it, not just the
explanation.

═════════════════════════════════════════════════════════════════
OUTPUT FORMAT — return ONLY valid JSON, no prose, no code fences.
═════════════════════════════════════════════════════════════════

The JSON has FOUR top-level sections (`phases`, `explanations`,
`disturbances`, and one key per rubric question). Do not skip any.

EVERY timestamp in your output uses the format:

    "WALLCLOCK (VIDEO_OFFSET)"

where:
  - WALLCLOCK    = the burned-in CCTV clock visible in the frame at
                   that moment, format `HH:MM:SS` (e.g. `09:18:42`).
                   Read it from the on-screen overlay.
  - VIDEO_OFFSET = elapsed time from the start of THIS video clip,
                   also `HH:MM:SS` (e.g. `00:05:25`).

Examples:
  "09:18:42 (00:05:25)"
  "10:34:11 (01:20:54)"

The wall clock is what the user will use to verify your answers — it
matches the centre's CCTV recording timestamps directly. If the wall
clock is illegible at a given moment, write `null` in its place:
`"null (00:05:25)"`. Do NOT omit the parenthesised video offset; both
parts are required so we can cross-check.

{{
  "phases": [
    {{
      "start":           "WALLCLOCK (VIDEO_OFFSET)",
      "end":             "WALLCLOCK (VIDEO_OFFSET)",
      "type":            "art_gym | art_games | artiverse_or_artistotle | experience_book | art_care | other",
      "what_happened":   "1 sentence summary of teacher + children activity",
      "children_present": true | false
    }},
    ...one entry per distinct block you observe...
  ],

  "explanations": [
    {{
      "ts":               "WALLCLOCK (VIDEO_OFFSET)",
      "activity":         "art_gym | art_games | artiverse_or_artistotle | other",
      "summary":          "what the teacher explained (1 sentence)",
      "was_clear":        "yes | no | partial",
      "confidence_tone":  "confident | hesitant | mixed",
      "children_engaged_after": "<approximate number of children who appeared to understand and start the activity>",
      "children_questions": [
        {{
          "ts":                          "WALLCLOCK (VIDEO_OFFSET)",
          "asker":                       "<child descriptor — clothing/position, or 'unidentified'>",
          "question":                    "<what the child asked (1 sentence)>",
          "teacher_response":            "<how the teacher answered (1 sentence)>",
          "response_addressed_question": "yes | partially | no"
        }},
        ...one entry per question a child asked during or just after this
        explanation; leave the array empty if no questions were asked...
      ]
    }},
    ...one entry per distinct explanation event the teacher gives...
  ],

  "disturbances": [
    {{
      "ts":               "WALLCLOCK (VIDEO_OFFSET)",
      "cause":            "<who or what caused the disturbance — describe the child by clothing/position if known, or 'multiple children', 'teacher', 'external'>",
      "description":      "<what the disturbance was (1 sentence)>",
      "teacher_response": "<what the teacher actually did in response (1 sentence)>",
      "resolution":       "ended | partially ended | persisted | not_addressed",
      "resolved_at":      "WALLCLOCK (VIDEO_OFFSET) or null if not resolved within the recording"
    }},
    ...one entry per distinct disturbance event in the class; leave the
    array empty if none observed...
  ],

  "Q1": {{
    "answer":     "...",
    "confidence": "high|medium|low",
    "evidence":   "WALLCLOCK (VIDEO_OFFSET) — what you observed; if multiple moments, separate them with ` ; `"
  }},
  "Q2": {{ "answer": "...", "confidence": "...", "evidence": "..." }},
  ...one entry per rubric question...
}}

═════════════════════════════════════════════════════════════════
HOW TO PROCEED
═════════════════════════════════════════════════════════════════

STEP 1 — Build the "phases" array.
  Watch the whole video and enumerate every distinct activity block.
  Use the phase types from the glossary. Multiple blocks of the same
  type are fine (e.g. two "transition" blocks). If a teacher RE-uses
  a phase later in the session, list it as a separate phase entry.

  **REQUIREMENT — full coverage**: your `phases` array MUST cover the
  ENTIRE video from `00:00:00` to the final timestamp. The `end` of
  your LAST phase entry must match (within ~30 sec) the timestamp at
  the very end of the video. If your phases stop short of that, you
  have not completed STEP 1 — go back and continue enumerating phases
  for the unaccounted-for time.

  **REQUIREMENT — phase ends at activity end, NOT explanation end**:
  if children continue an art game / gym / artwork after the teacher
  finishes explaining, the phase continues until the children pack up
  and move on. See the CRITICAL section in the glossary above.

  **SELF-CHECK at end of STEP 1** (do this before writing your output):
    1. Does the FIRST phase's `start` ≈ `{wallclock_start} (00:00:00)`?
    2. Does the LAST phase's `end` ≈ `{wallclock_end} ({duration_str})` (within 30 sec)?
       **The video runs for {duration_str} total
       (CCTV wall clock: {wallclock_start} → {wallclock_end} IST).
       If your last phase ends before that, you are missing time. Most
       likely the children continued Artiverse/Artistotle, did
       Experience Book, did Art Care, or there is post-session time you
       should mark as `other`.**
    3. Sum every phase's duration: does it equal `{duration_sec}` seconds?
    4. Are there NO gaps between consecutive phases (one ends where
       the next begins)?
    5. Did you account for any unstructured / off-segment time during
       the session? Mark such blocks as `other`.
  If ANY of these fail, REDO STEP 1 — extend or add phases until the
  whole video is covered.

STEP 2 — Build the "explanations" array.
  The teacher will often explain the same activity multiple times
  (initial group explanation, table-by-table re-explanation,
  individual re-explanation). Capture EVERY distinct explanation
  event, even if it covers the same activity twice. This is critical
  for the Content Knowledge questions.

  For each explanation, also capture any questions children asked
  during or just after that explanation, and the teacher's actual
  response to each. Put these into the explanation entry's
  `children_questions` sub-array. If no child asked anything for an
  explanation, leave that sub-array empty.

STEP 3 — Build the "disturbances" array.
  A disturbance is any moment that disrupted the flow of teaching or
  the children's engagement. Causes can be:
    - one child (describe by clothing / position / known role)
    - multiple children
    - the teacher (e.g. left the room, missed a moment)
    - external (someone enters, a noise from outside, materials
      spilled, etc.)
  For EACH disturbance, capture: when it started (`ts`), what caused
  it, what the disturbance was, what the teacher did in response,
  whether it ended after the teacher's intervention (`resolution`),
  and the moment it resolved (`resolved_at`). Leave the array empty
  if no disturbance was observed.

STEP 4 — Answer each rubric question using your phase, explanation,
  and disturbance data as ground truth.

═════════════════════════════════════════════════════════════════
CRITICAL RULES
═════════════════════════════════════════════════════════════════

1. WATCH THE WHOLE VIDEO. Sample timestamps from beginning, middle, and end.

2. TIMESTAMPS REQUIRED. Every Q answer cites at least one HH:MM:SS timestamp.
   For pattern questions ("how many disruptions", "did the teacher manage..."),
   cite MULTIPLE timestamps from different parts of the video.

3. ANCHOR ANSWERS IN PHASES.
   - For Q5–Q9 (minutes per segment): sum durations from your "phases" array.
     The rubric asks for minutes spent on Art Games (Q5), Art Gym (Q6),
     Artiverse/Artistotle (Q7), Experience Book (Q8), and Art Care (Q9).
     Per the Openhouse programme, Art Gym IS the warm-up — do not duplicate
     it under a separate "warm-up" count. If a segment truly did not occur,
     answer 0 — but only if it truly did not occur, NOT because you weren't
     paying attention to it.
   - For Q10–Q13 (did teacher explain games / gym correctly + confidently):
     judge from your "explanations" array, integrating ALL explanation events
     for that activity. If the teacher explained the art game 3 times, judge
     the OVERALL quality across all three.
   - For Q24–Q26 (disruptions, how handled, ended after intervention):
     use your "disturbances" array as ground truth.
       · Q24 = `len(disturbances)`.
       · Q25 = describe the teacher_response field across all entries.
       · Q26 = roll up the resolution field across all entries (e.g.
         "all ended" / "2 of 3 ended" / "persisted").

4. WHEN TO USE "INSUFFICIENT INFORMATION":
   - The relevant phase/event TRULY did not occur (verify by checking your
     "phases" array — if the phase isn't there at all, then yes, insufficient)
   - The relevant moment is off-camera, inaudible, or muted
   - DO NOT use it just because the question is subjective.

5. ANALYSIS TAG GUIDANCE — each question has [brackets]:
     [Visual]        — answer purely from what you see
     [Audio]         — answer purely from what you hear
     [Audio - tone]  — answer based on vocal tone/affect
     [Visual + Audio]— integrate both
   If the audio is degraded and the question is [Audio]/[Audio - tone],
   say so explicitly and lower confidence — but still try to answer.

6. NO GUESSING. Do not invent names, dialogue, or events you didn't observe.

7. NUMERIC ANSWERS. Where the question asks for a count or duration,
   start the answer with the number ("3 children — ..." or "approximately
   12 minutes — ...") not a long description.

═════════════════════════════════════════════════════════════════
QUESTIONS, GROUPED BY CRITERIA
═════════════════════════════════════════════════════════════════

{questions_block}
"""


def format_questions_block(questions: list[dict]) -> str:
    lines = []
    current_criteria = None
    current_group = None
    for q in questions:
        if q["criteria"] != current_criteria:
            current_criteria = q["criteria"]
            current_group = None
            lines.append(f"\n=== {current_criteria} ===")
        if q["group"] != current_group:
            current_group = q["group"]
            lines.append(f"\n[Group] {current_group}")
        analysis_tag = f"[{q['analysis']}]" if q['analysis'] else ""
        input_hint = f"  (input ref: {q['input_required']})" if q['input_required'] else ""
        lines.append(f"  {q['id']} {analysis_tag}: {q['question']}{input_hint}")
    return "\n".join(lines)


def stage4_rubric(trimmed: Path, questions: list[dict], llm: LLMAdapter,
                  model: str, run_dir: Path, force: bool = False) -> dict:
    answers_path = run_dir / "4_rubric_answers.json"
    if answers_path.exists() and not force:
        log.info(f"[4] rubric: {answers_path.name} exists, skipping")
        return json.loads(answers_path.read_text())

    # Inject the trimmed video's actual duration so Gemini knows how far its
    # phase enumeration must reach.
    dur_sec_f = video_duration_seconds(trimmed)
    dur_sec = int(round(dur_sec_f))
    h, rem = divmod(dur_sec, 3600)
    m, s = divmod(rem, 60)
    duration_str = f"{h:02d}:{m:02d}:{s:02d}"

    # Derive the trimmed video's wall-clock anchors from the boundaries.json
    # (video_start_wall_clock of the COMBINED file + the trim's start offset).
    # Falls back to gracefully-templated placeholders if anything is missing.
    boundaries_path = run_dir / "2_boundaries.json"
    wallclock_start = "??:??:??"
    wallclock_end = "??:??:??"
    if boundaries_path.exists():
        try:
            b = json.loads(boundaries_path.read_text())
            combined_start_wc = b.get("video_start_wall_clock")
            # Re-derive trim window from boundaries to get our start_sec
            comb_dur = video_duration_seconds(run_dir / "1a_combined.mp4")
            trim_start_sec, _ = compute_trim_window(b, comb_dur)
            if combined_start_wc:
                base_sec = _parse_hms(combined_start_wc) or 0
                start_total = base_sec + int(trim_start_sec)
                end_total = start_total + dur_sec
                def _fmt(total_sec: int) -> str:
                    h = (total_sec // 3600) % 24
                    m = (total_sec // 60) % 60
                    s = total_sec % 60
                    return f"{h:02d}:{m:02d}:{s:02d}"
                wallclock_start = _fmt(start_total)
                wallclock_end = _fmt(end_total)
        except Exception as e:
            log.warning(f"[4] couldn't derive wall-clock anchors: {e}")

    prompt = RUBRIC_PROMPT_TEMPLATE.format(
        questions_block=format_questions_block(questions),
        duration_str=duration_str,
        duration_sec=dur_sec,
        wallclock_start=wallclock_start,
        wallclock_end=wallclock_end,
    )
    (run_dir / "4_rubric_prompt.txt").write_text(prompt)
    log.info(f"[4] rubric: video duration {duration_str} ({dur_sec}s)")
    log.info(f"[4] rubric: wall-clock anchor {wallclock_start} → {wallclock_end} IST")
    log.info(f"[4] rubric: uploading {trimmed.name} + asking {model} about {len(questions)} questions")

    video_file = llm.upload_video(trimmed)

    # Per-model thinking_budget (Gemini 2.5 disables thinking; 3.x requires it)
    if model.startswith("gemini-2.5"):
        thinking_budget = 0
        max_output_tokens = 30000
    else:
        thinking_budget = None
        max_output_tokens = 65536

    t0 = time.time()
    raw = llm.call_gemini_video(
        prompt=prompt,
        video_file=video_file,
        model_name=model,
        max_output_tokens=max_output_tokens,
        temperature=0.0,
        force_json=True,
        thinking_budget=thinking_budget,
        fps=RUBRIC_FPS,
    )
    log.info(f"[4] rubric: response in {time.time()-t0:.0f}s ({len(raw)} chars)")
    (run_dir / "4_rubric_raw.txt").write_text(raw)

    parsed = parse_json_lenient(raw)
    if not isinstance(parsed, dict):
        log.error(f"[4] rubric: parse failed; expected dict, got {type(parsed).__name__}")
        return {}

    # Pull the new top-level arrays into their own files so stage 5's CSV
    # stays focused on rubric questions, while the structured side outputs
    # are preserved for verification.
    phases = parsed.pop("phases", []) or []
    explanations = parsed.pop("explanations", []) or []
    disturbances = parsed.pop("disturbances", []) or []
    (run_dir / "4_phases.json").write_text(json.dumps(phases, indent=2, ensure_ascii=False))
    (run_dir / "4_explanations.json").write_text(json.dumps(explanations, indent=2, ensure_ascii=False))
    (run_dir / "4_disturbances.json").write_text(json.dumps(disturbances, indent=2, ensure_ascii=False))
    total_qs = sum(len(e.get("children_questions") or []) for e in explanations)
    log.info(
        f"[4] rubric: extracted {len(phases)} phase(s) + {len(explanations)} explanation event(s) "
        f"(+{total_qs} child question(s)) + {len(disturbances)} disturbance(s)"
    )
    if phases:
        log.info("[4] phases detected:")
        for p in phases:
            children = "✓" if p.get("children_present") else "✗"
            log.info(f"    {p.get('start','??')}–{p.get('end','??')}  "
                     f"[{p.get('type','?'):<23}] kids={children}  "
                     f"— {p.get('what_happened','')}")
    if explanations:
        log.info("[4] explanation events:")
        for e in explanations:
            log.info(f"    {e.get('ts','??')}  [{e.get('activity','?'):<23}] "
                     f"clear={e.get('was_clear','?'):<7} tone={e.get('confidence_tone','?'):<9} "
                     f"engaged={e.get('children_engaged_after','?')}  — {e.get('summary','')}")
            for q in (e.get("children_questions") or []):
                log.info(f"      Q@{q.get('ts','??')}  asker={q.get('asker','?')}  "
                         f"addressed={q.get('response_addressed_question','?')}  "
                         f"— {q.get('question','')} → {q.get('teacher_response','')}")
    if disturbances:
        log.info("[4] disturbances:")
        for d in disturbances:
            log.info(f"    {d.get('ts','??')}  cause={d.get('cause','?')}  "
                     f"resolution={d.get('resolution','?'):<18} "
                     f"resolved_at={d.get('resolved_at','?')}")
            log.info(f"      desc: {d.get('description','')}")
            log.info(f"      teacher: {d.get('teacher_response','')}")

    # Keep only Q* entries in the rubric answers file
    answers = {str(k): v for k, v in parsed.items() if str(k).startswith("Q")}
    answers_path.write_text(json.dumps(answers, indent=2, ensure_ascii=False))
    log.info(f"[4] rubric: parsed {len(answers)} answers")
    return answers


# ─── Stage 5: report ────────────────────────────────────────────────────────

def stage5_report(questions: list[dict], answers: dict, run_dir: Path) -> None:
    csv_path = run_dir / "5_report.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Q ID", "Criteria", "Group", "Question", "Analysis",
                    "Answer", "Confidence", "Evidence", "Status"])
        answered = 0
        insufficient = 0
        missing = 0
        for q in questions:
            a = answers.get(q["id"]) or {}
            ans_text = ""
            conf = ""
            ev = ""
            status = "missing"
            if isinstance(a, dict):
                ans_text = str(a.get("answer", "")).strip()
                conf = str(a.get("confidence", "")).strip()
                ev = str(a.get("evidence", "")).strip()
                if not ans_text:
                    status = "missing"
                    missing += 1
                elif ans_text.upper().startswith("INSUFFICIENT"):
                    status = "insufficient"
                    insufficient += 1
                else:
                    status = "answered"
                    answered += 1
            else:
                missing += 1
            w.writerow([q["id"], q["criteria"], q["group"], q["question"],
                        q["analysis"], ans_text, conf, ev, status])

    log.info("")
    log.info(f"[5] report: {csv_path}")
    log.info(f"[5] summary: answered={answered}  insufficient_info={insufficient}  missing={missing}  total={len(questions)}")

    # Print console-friendly summary by criteria
    log.info("")
    log.info("Per-criteria breakdown:")
    by_crit: dict[str, dict[str, int]] = {}
    for q in questions:
        crit = q["criteria"]
        a = answers.get(q["id"]) or {}
        bucket = by_crit.setdefault(crit, {"answered": 0, "insufficient": 0, "missing": 0, "total": 0})
        bucket["total"] += 1
        if not isinstance(a, dict) or not a.get("answer"):
            bucket["missing"] += 1
        elif str(a["answer"]).upper().startswith("INSUFFICIENT"):
            bucket["insufficient"] += 1
        else:
            bucket["answered"] += 1
    for crit, b in by_crit.items():
        log.info(f"  {crit:<25} answered={b['answered']:>3}  insufficient={b['insufficient']:>3}  missing={b['missing']:>3}  (of {b['total']})")


# ─── Main ───────────────────────────────────────────────────────────────────

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[1],
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--rubric", required=True, type=Path,
                   help="Path to the rubric .xlsx")
    p.add_argument("--sheet", default="Art",
                   help="Sheet name within the rubric .xlsx (default: Art)")
    p.add_argument("--segments", required=True, nargs="+", type=Path,
                   help="Raw NVR mp4 segment(s) to combine (in chronological order)")
    p.add_argument("--model", default="gemini-3.1-pro-preview",
                   help="Gemini model for the rubric pass (default: gemini-3.1-pro-preview)")
    p.add_argument("--out-dir", type=Path,
                   default=ROOT / "data" / "art_rubric_runs" / datetime.now().strftime("%Y-%m-%d_%H%M%S"),
                   help="Output dir (default: data/art_rubric_runs/<timestamp>)")
    p.add_argument("--force-from", choices=["combine", "boundaries", "trim", "rubric", "report"],
                   help="Re-run from this stage onward (clearing later cached outputs)")
    args = p.parse_args()

    if not args.rubric.exists():
        log.error(f"Rubric file not found: {args.rubric}")
        return 1

    run_dir: Path = args.out_dir
    run_dir.mkdir(parents=True, exist_ok=True)
    log.info(f"Output dir: {run_dir}")

    # Force semantics: if --force-from=trim, we re-do trim + rubric + report
    stages = ["combine", "boundaries", "trim", "rubric", "report"]
    force_idx = stages.index(args.force_from) if args.force_from else len(stages)
    def f(stage: str) -> bool:
        return stages.index(stage) >= force_idx

    # Load + log the rubric
    questions = parse_rubric_sheet(args.rubric, args.sheet)
    log.info(f"Loaded {len(questions)} rubric questions from sheet {args.sheet!r}")
    by_crit_count: dict[str, int] = {}
    for q in questions:
        by_crit_count[q["criteria"]] = by_crit_count.get(q["criteria"], 0) + 1
    for c, n in by_crit_count.items():
        log.info(f"  {c}: {n} question(s)")

    combined = run_dir / "1a_combined.mp4"
    boundary_input = run_dir / "1b_boundary_input.mp4"
    trimmed = run_dir / "3_trimmed.mp4"

    # Derive wall-clock anchor from the FIRST segment's filename. The NVR
    # naming convention encodes the recording start time as HHMMSS in the
    # filename; this is our durable anchor in case Gemini misreads the
    # burned-in clock at frame 0.
    first_segment_anchor = derive_wall_clock_from_filename(args.segments[0])
    if first_segment_anchor:
        log.info(f"Wall-clock anchor from first segment filename: {first_segment_anchor}")
    else:
        log.warning(
            f"Could not parse wall-clock anchor from filename: {args.segments[0].name}. "
            "Boundary detection will rely solely on Gemini's frame-0 reading."
        )

    overall_t0 = time.time()
    try:
        stage1_combine(args.segments, combined, boundary_input, force=f("combine"))
        boundaries = stage2_boundaries(
            boundary_input, combined, run_dir, LLMAdapter(),
            filename_anchor=first_segment_anchor,
            force=f("boundaries"),
        )
        dur = video_duration_seconds(combined)
        start_sec, end_sec = compute_trim_window(boundaries, dur)
        stage3_trim(combined, start_sec, end_sec, trimmed, force=f("trim"))
        answers = stage4_rubric(trimmed, questions, LLMAdapter(), args.model, run_dir, force=f("rubric"))
        stage5_report(questions, answers, run_dir)
    except Exception as e:
        log.error(f"Aborted: {type(e).__name__}: {e}")
        raise

    log.info("")
    log.info(f"Done in {(time.time()-overall_t0)/60:.1f} min")
    return 0


if __name__ == "__main__":
    sys.exit(main())
