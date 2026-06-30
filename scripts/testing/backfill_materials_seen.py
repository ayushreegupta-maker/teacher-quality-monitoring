"""
One-off backfill: walk data/rubric_runs/<subject>/<run>/5_answers.json,
extract materials_seen, and write three columns into the Runs sheet of
data/tqm_answers.xlsx so the cloud dashboard can read them without the
per-run JSON / session directory being available:

  - materials_seen_json           — Shape B reasoner's materials list
  - trimmed_video_duration_seconds — ffprobe of 3_trimmed.mp4
  - boundaries_detected           — whether 2_boundaries.json exists

Idempotent: re-running won't duplicate columns and updates in-place.

Run once after each schema change:
    .venv/bin/python scripts/testing/backfill_materials_seen.py
"""
from __future__ import annotations

import json
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
XLSX = ROOT / "data" / "tqm_answers.xlsx"
RUBRIC_RUNS = ROOT / "data" / "rubric_runs"
SESSIONS_DIR = ROOT / "data" / "sessions"

NEW_COLUMNS = [
    "materials_seen_json",
    "trimmed_video_duration_seconds",
    "boundaries_detected",
]


def ffprobe_duration(path: Path) -> float | None:
    try:
        r = subprocess.run(
            ["ffprobe", "-v", "error", "-show_format",
             "-print_format", "json", str(path)],
            capture_output=True, text=True, timeout=15,
        )
        return float(json.loads(r.stdout)["format"]["duration"])
    except Exception:
        return None


def materials_index_from_disk() -> dict[str, str]:
    """{config_slug → json-encoded materials_seen list}."""
    out: dict[str, str] = {}
    if not RUBRIC_RUNS.exists():
        return out
    for subj_dir in RUBRIC_RUNS.iterdir():
        if not subj_dir.is_dir():
            continue
        for run_dir in subj_dir.iterdir():
            if not run_dir.is_dir():
                continue
            p = run_dir / "5_answers.json"
            if not p.exists():
                continue
            try:
                data = json.loads(p.read_text())
            except Exception as e:
                print(f"  warn: bad JSON at {p.relative_to(ROOT)}: {e}")
                continue
            ms = data.get("materials_seen")
            if not ms:
                continue
            out[run_dir.name] = json.dumps(ms)
    return out


def session_window_info(subject: str, session_id: str) -> tuple[float | None, bool]:
    """Return (trimmed_duration_seconds_or_None, boundaries_detected_bool)
    for a session, by reading 3_trimmed.mp4 + 2_boundaries.json from disk."""
    sess_dir = SESSIONS_DIR / subject / session_id
    trim = sess_dir / "3_trimmed.mp4"
    boundaries = sess_dir / "2_boundaries.json"
    dur = ffprobe_duration(trim) if trim.exists() else None
    return dur, boundaries.exists()


def main():
    if not XLSX.exists():
        print(f"missing: {XLSX}")
        sys.exit(1)

    backup = XLSX.with_name(
        f"tqm_answers.backup_{datetime.now().strftime('%Y%m%dT%H%M%S')}.xlsx"
    )
    shutil.copy2(XLSX, backup)
    print(f"backup → {backup.name}")

    materials_by_slug = materials_index_from_disk()
    print(f"loaded materials_seen for {len(materials_by_slug)} run(s) from disk")

    wb = openpyxl.load_workbook(XLSX)
    if "Runs" not in wb.sheetnames:
        print("ERROR: 'Runs' sheet missing from workbook")
        sys.exit(1)
    ws = wb["Runs"]

    headers = [c.value for c in ws[1]]
    for col in NEW_COLUMNS:
        if col not in headers:
            new_col = ws.max_column + 1
            ws.cell(row=1, column=new_col, value=col)
            headers.append(col)
            print(f"added column {col!r} at position {new_col}")

    slug_idx = headers.index("config_slug")
    subj_idx = headers.index("subject")
    sid_idx = headers.index("session_id")
    ms_idx = headers.index("materials_seen_json")
    dur_idx = headers.index("trimmed_video_duration_seconds")
    bd_idx = headers.index("boundaries_detected")

    n_materials = n_window = n_total = 0
    # cache session_window lookups so we don't ffprobe the same trim twice
    win_cache: dict[tuple[str, str], tuple[float | None, bool]] = {}

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column):
        n_total += 1
        slug = (row[slug_idx].value or "").strip() if row[slug_idx].value else ""
        subject = (row[subj_idx].value or "").strip() if row[subj_idx].value else ""
        sid = (row[sid_idx].value or "").strip() if row[sid_idx].value else ""

        if slug:
            ms = materials_by_slug.get(slug)
            if ms:
                ws.cell(row=row[0].row, column=ms_idx + 1, value=ms)
                n_materials += 1

        if subject and sid:
            key = (subject, sid)
            if key not in win_cache:
                win_cache[key] = session_window_info(subject, sid)
            dur, bd = win_cache[key]
            if dur is not None:
                ws.cell(row=row[0].row, column=dur_idx + 1, value=dur)
                n_window += 1
            ws.cell(row=row[0].row, column=bd_idx + 1, value=bool(bd))

    wb.save(XLSX)
    print(f"patched {n_total} Runs rows — "
          f"{n_materials} with materials, {n_window} with duration")


if __name__ == "__main__":
    main()
