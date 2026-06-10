"""
CCTV downloader — pulls classroom video from a Hikvision NVR via ISAPI over Tailscale.

Reads camera config from `data/cctv_cameras.xlsx`, searches for recordings on a
given date and time window, downloads each segment, transcodes to H.264 + AAC if
needed (so Gemini accepts it), and saves into `data/raw/`.

Usage:
  # Pull a single camera for a specific date + time window
  python scripts/cctv_pull.py --date 2026-05-30 --camera D29 \
      --start 09:00 --end 13:00

  # Pull ALL active cameras for a date (uses default window 09:00–17:00)
  python scripts/cctv_pull.py --date 2026-05-30 --all

  # Override default window
  python scripts/cctv_pull.py --date 2026-05-30 --all --start 09:00 --end 11:00

  # Dry run — search only, don't download
  python scripts/cctv_pull.py --date 2026-05-30 --camera D14 --dry-run

  # Skip transcoding (keep raw NVR output; smaller pipeline but not Gemini-ready)
  python scripts/cctv_pull.py --date 2026-05-30 --camera D14 --no-transcode

Output naming:
  data/raw/{camera_id}_{centre}_{subject}_{YYYYMMDD}_{startHHMMSS}.mp4

Requires (in .env at project root):
  NVR_USER=admin
  NVR_PASSWORD=<the device admin password>
"""

import argparse
import logging
import os
import re
import shutil
import subprocess
import sys
import time
import uuid
import xml.sax.saxutils as saxutils
from datetime import datetime, date as date_cls, timedelta
from pathlib import Path
from typing import Optional
from xml.etree import ElementTree as ET

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl
import requests
from requests.auth import HTTPDigestAuth
from dotenv import load_dotenv

load_dotenv(ROOT / ".env", override=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("cctv_pull")
logging.getLogger("urllib3").setLevel(logging.WARNING)


# ─── Paths & config ────────────────────────────────────────────────────────

CAMERA_CONFIG = ROOT / "data" / "cctv_cameras.xlsx"
OUTPUT_DIR = ROOT / "data" / "raw"
TMP_DIR = ROOT / "data" / "raw" / ".tmp"

HIK_NS = {"hik": "http://www.hikvision.com/ver20/XMLSchema"}
IST_OFFSET = "+05:30"


# ─── Helpers ───────────────────────────────────────────────────────────────

def channel_from_camera_id(camera_id: str) -> int:
    """`D14` → `14`. Accepts `D14`, `d14`, `14`."""
    m = re.match(r"[Dd]?(\d+)$", str(camera_id).strip())
    if not m:
        raise ValueError(f"Cannot parse channel from camera_id={camera_id!r}")
    return int(m.group(1))


def track_id_main(channel: int) -> int:
    """Hikvision convention: trackID = channel*100 + 1 (main stream)."""
    return channel * 100 + 1


def load_cameras(only: Optional[set[str]] = None) -> list[dict]:
    """Read the Excel sheet. Returns a list of dicts, filtered by is_active and
    optionally by `only` (set of camera_ids)."""
    if not CAMERA_CONFIG.exists():
        log.error(f"Camera config not found: {CAMERA_CONFIG}")
        sys.exit(1)

    wb = openpyxl.load_workbook(CAMERA_CONFIG, data_only=True)
    ws = wb["cameras"] if "cameras" in wb.sheetnames else wb.active
    headers = [c.value for c in ws[1]]
    cams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rec = dict(zip(headers, row))
        active = rec.get("is_active")
        if not (active is True or str(active).strip().lower() in ("true", "1", "yes", "y")):
            continue
        if only is not None and rec["camera_id"] not in only:
            continue
        cams.append(rec)
    return cams


def parse_hik_time(s: str) -> datetime:
    """Hik returns 2026-05-30T08:31:41Z. Despite the 'Z', it's actually the NVR's
    local time (IST). Strip the Z and treat as naive."""
    return datetime.strptime(s.replace("Z", ""), "%Y-%m-%dT%H:%M:%S")


def fmt_filename_safe(s: str) -> str:
    """Sanitise a string for use in filenames."""
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(s)).strip("_").lower()


# ─── ISAPI calls ───────────────────────────────────────────────────────────

def isapi_search(host: str, user: str, password: str, track_id: int,
                 start_dt: datetime, end_dt: datetime, max_results: int = 50) -> list[dict]:
    """Search recordings on a track within a time window. Returns segments."""
    payload = f"""<?xml version="1.0" encoding="UTF-8"?>
<CMSearchDescription xmlns="http://www.hikvision.com/ver20/XMLSchema">
  <searchID>{uuid.uuid4()}</searchID>
  <trackList><trackID>{track_id}</trackID></trackList>
  <timeSpanList>
    <timeSpan>
      <startTime>{start_dt.strftime('%Y-%m-%dT%H:%M:%S')}{IST_OFFSET}</startTime>
      <endTime>{end_dt.strftime('%Y-%m-%dT%H:%M:%S')}{IST_OFFSET}</endTime>
    </timeSpan>
  </timeSpanList>
  <maxResults>{max_results}</maxResults>
  <searchResultPostion>0</searchResultPostion>
</CMSearchDescription>"""

    r = requests.post(
        f"http://{host}/ISAPI/ContentMgmt/search",
        data=payload.encode("utf-8"),
        auth=HTTPDigestAuth(user, password),
        headers={"Content-Type": "application/xml"},
        timeout=30,
    )
    r.raise_for_status()

    root = ET.fromstring(r.text)
    # Check status — Hik sometimes returns 200 with an error body
    status_node = root.find("hik:responseStatus", HIK_NS)
    if status_node is not None and status_node.text != "true":
        msg = root.findtext("hik:responseStatusStrg", default="(no message)", namespaces=HIK_NS)
        log.warning(f"Search returned non-OK status: {msg}")
        return []

    segs = []
    for item in root.findall("hik:matchList/hik:searchMatchItem", HIK_NS):
        start = item.findtext("hik:timeSpan/hik:startTime", namespaces=HIK_NS)
        end = item.findtext("hik:timeSpan/hik:endTime", namespaces=HIK_NS)
        uri = item.findtext("hik:mediaSegmentDescriptor/hik:playbackURI",
                            namespaces=HIK_NS)
        if not uri:
            continue
        segs.append({
            "start": parse_hik_time(start),
            "end": parse_hik_time(end),
            "playback_uri": uri,
        })
    return segs


def isapi_download(host: str, user: str, password: str,
                   playback_uri: str, out_path: Path,
                   timeout_connect: int = 30, timeout_read: int = 1800) -> int:
    """Download one segment by playbackURI. Returns total bytes written."""
    # The URI from ET has `&` (unescaped); the request body needs `&amp;`.
    escaped = saxutils.escape(playback_uri)
    payload = f"""<?xml version="1.0" encoding="UTF-8"?>
<downloadRequest xmlns="http://www.hikvision.com/ver20/XMLSchema">
  <playbackURI>{escaped}</playbackURI>
</downloadRequest>"""

    r = requests.post(
        f"http://{host}/ISAPI/ContentMgmt/download",
        data=payload.encode("utf-8"),
        auth=HTTPDigestAuth(user, password),
        headers={"Content-Type": "application/xml"},
        stream=True,
        timeout=(timeout_connect, timeout_read),
    )
    r.raise_for_status()

    # Pull expected segment size from the playbackURI's `&size=...` param for
    # % progress. URI looks like: rtsp://.../?starttime=...&endtime=...&size=1063100828
    m = re.search(r"[?&]size=(\d+)", playback_uri)
    expected_total = int(m.group(1)) if m else None

    LOG_INTERVAL_SEC = 300   # log progress every 5 min, not every 5 sec
    out_path.parent.mkdir(parents=True, exist_ok=True)
    total = 0
    t0 = time.time()
    last_log = t0
    with open(out_path, "wb") as f:
        for chunk in r.iter_content(chunk_size=64 * 1024):
            if chunk:
                f.write(chunk)
                total += len(chunk)
                now = time.time()
                if now - last_log >= LOG_INTERVAL_SEC:
                    elapsed = now - t0
                    speed_mbps = total / elapsed / 1e6
                    if expected_total:
                        pct = total / expected_total * 100
                        eta_sec = (expected_total - total) / max(total / elapsed, 1)
                        log.info(
                            f"    … {total/1e6:.0f}/{expected_total/1e6:.0f} MB "
                            f"({pct:.0f}%) in {elapsed/60:.1f} min "
                            f"@ {speed_mbps:.2f} MB/s, ETA {eta_sec/60:.1f} min"
                        )
                    else:
                        log.info(
                            f"    … {total/1e6:.0f} MB in {elapsed/60:.1f} min "
                            f"@ {speed_mbps:.2f} MB/s"
                        )
                    last_log = now
    return total


# ─── Transcoding ───────────────────────────────────────────────────────────

def probe_codecs(path: Path) -> tuple[str, str]:
    """Returns (video_codec, audio_codec). Either may be '' if not present."""
    def probe(stream: str) -> str:
        try:
            out = subprocess.check_output(
                ["ffprobe", "-v", "error",
                 "-select_streams", stream,
                 "-show_entries", "stream=codec_name",
                 "-of", "csv=p=0", str(path)],
                stderr=subprocess.DEVNULL,
            ).decode().strip()
            return out
        except subprocess.CalledProcessError:
            return ""
    return probe("v:0"), probe("a:0")


def transcode_to_h264_aac(in_path: Path, out_path: Path) -> None:
    log.info(f"    transcoding → {out_path.name}")
    subprocess.run([
        "ffmpeg", "-y", "-i", str(in_path),
        "-c:v", "libx264", "-crf", "23", "-preset", "fast",
        "-c:a", "aac", "-b:a", "96k",
        "-movflags", "+faststart",
        str(out_path),
    ], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def ensure_gemini_ready(raw_path: Path, final_path: Path) -> None:
    """If raw_path is already h264+aac, just rename. Otherwise transcode."""
    v, a = probe_codecs(raw_path)
    log.info(f"    raw codecs: video={v!r} audio={a!r}")
    if v == "h264" and a in ("aac", ""):
        raw_path.rename(final_path)
    else:
        transcode_to_h264_aac(raw_path, final_path)
        raw_path.unlink(missing_ok=True)


# ─── Naming ────────────────────────────────────────────────────────────────

def output_filename(cam: dict, segment_start: datetime) -> str:
    centre = fmt_filename_safe(cam["centre_name"])
    subject = fmt_filename_safe(cam["subject"])
    date_str = segment_start.strftime("%Y%m%d")
    time_str = segment_start.strftime("%H%M%S")
    return f"{cam['camera_id']}_{centre}_{subject}_{date_str}_{time_str}.mp4"


# ─── Main per-camera flow ──────────────────────────────────────────────────

def pull_camera(cam: dict, date: date_cls, start_hm: str, end_hm: str,
                user: str, password: str, do_transcode: bool, dry_run: bool) -> dict:
    """Search and download a single camera's segments for the given window."""
    cam_id = cam["camera_id"]
    host = cam["nvr_host"]
    channel = channel_from_camera_id(cam_id)
    track_id = track_id_main(channel)

    start_dt = datetime.combine(date, datetime.strptime(start_hm, "%H:%M").time())
    end_dt = datetime.combine(date, datetime.strptime(end_hm, "%H:%M").time())

    log.info(f"━━ {cam_id} @ {cam['centre_name']} / {cam['subject']} ━━")
    log.info(f"  window {start_dt} → {end_dt} (IST)  track={track_id}  host={host}")

    try:
        segs = isapi_search(host, user, password, track_id, start_dt, end_dt)
    except Exception as e:
        log.error(f"  search failed: {type(e).__name__}: {e}")
        return {"camera_id": cam_id, "error": str(e), "segments": 0, "downloaded": 0}

    log.info(f"  {len(segs)} segment(s) returned")
    if dry_run:
        for s in segs:
            log.info(f"    {s['start']} → {s['end']}")
        return {"camera_id": cam_id, "segments": len(segs), "downloaded": 0,
                "dry_run": True}

    downloaded = 0
    skipped = 0
    failed = 0
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for i, seg in enumerate(segs, 1):
        final_name = output_filename(cam, seg["start"])
        final_path = OUTPUT_DIR / final_name
        if final_path.exists():
            log.info(f"  [{i}/{len(segs)}] {final_name} already exists — skip")
            skipped += 1
            continue

        raw_path = TMP_DIR / (final_name.removesuffix(".mp4") + ".raw.mp4")
        log.info(f"  [{i}/{len(segs)}] {seg['start']} → {seg['end']}  → {final_name}")

        t0 = time.time()
        try:
            n = isapi_download(host, user, password, seg["playback_uri"], raw_path)
            log.info(f"    downloaded {n/1e6:.0f} MB in {time.time()-t0:.0f}s")
        except Exception as e:
            log.error(f"    download failed: {type(e).__name__}: {e}")
            raw_path.unlink(missing_ok=True)
            failed += 1
            continue

        try:
            if do_transcode:
                ensure_gemini_ready(raw_path, final_path)
            else:
                raw_path.rename(final_path)
            downloaded += 1
        except Exception as e:
            log.error(f"    post-processing failed: {type(e).__name__}: {e}")
            failed += 1

    return {"camera_id": cam_id, "segments": len(segs),
            "downloaded": downloaded, "skipped": skipped, "failed": failed}


# ─── CLI ───────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__.splitlines()[1],
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--date", required=True, type=lambda s: date_cls.fromisoformat(s),
                   help="Date to pull, YYYY-MM-DD (IST)")
    target = p.add_mutually_exclusive_group(required=True)
    target.add_argument("--camera", help="Single camera_id, e.g. D14")
    target.add_argument("--all", action="store_true", help="All active cameras")
    p.add_argument("--start", default="09:00", help="Window start HH:MM (IST). Default 09:00")
    p.add_argument("--end", default="17:00", help="Window end HH:MM (IST). Default 17:00")
    p.add_argument("--no-transcode", action="store_true",
                   help="Skip H.264/AAC transcode (faster but not Gemini-ready)")
    p.add_argument("--dry-run", action="store_true",
                   help="Search only; list segments without downloading")
    return p.parse_args()


def main() -> int:
    args = parse_args()

    user = os.environ.get("NVR_USER")
    password = os.environ.get("NVR_PASSWORD")
    if not user or not password:
        log.error("NVR_USER / NVR_PASSWORD missing. Set them in .env at the project root.")
        return 1

    cams = load_cameras(only={args.camera} if args.camera else None)
    if not cams:
        log.error("No matching active cameras in data/cctv_cameras.xlsx")
        return 1

    log.info(f"Will process {len(cams)} camera(s) for {args.date}")
    results = []
    t0 = time.time()
    for cam in cams:
        results.append(pull_camera(
            cam, args.date, args.start, args.end,
            user, password, not args.no_transcode, args.dry_run,
        ))

    log.info("")
    log.info(f"Done in {(time.time()-t0)/60:.1f} min")
    for r in results:
        log.info(f"  {r['camera_id']}: segments={r.get('segments', 0)} "
                 f"downloaded={r.get('downloaded', 0)} skipped={r.get('skipped', 0)} "
                 f"failed={r.get('failed', 0)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
