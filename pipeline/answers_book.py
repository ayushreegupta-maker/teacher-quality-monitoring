"""
The rolling answer accumulator: data/tqm_answers.xlsx.

One workbook, five tabs (Art / Public Speaking / Robotics / Runs / README),
appended-to by every `scripts/run_rubric.py` invocation. See PLAN.md §3.3.

Public API
----------
    init_workbook(xlsx_path)
        Build a fresh empty workbook with all 5 tabs + headers + README text.
        Idempotent — no-op if xlsx_path already exists.

    write_sidecar(queue_dir, *, answer_set, rubric, config, run_id, started_at,
                  finished_at, wall_clock_seconds, config_slug, run_n)
        -> Path

        Emit a single self-contained sidecar JSON to data/_answer_queue/.
        Contains exactly the rows that need to be appended to the subject
        tab + the Runs tab. Sidecars are the durable atomic write — even
        if the XLSX merge fails, the answers are safe on disk.

    merge_queue(xlsx_path, queue_dir, *, backup_path=None) -> dict
        Atomic merge protocol:
          1. cp xlsx_path -> backup_path (default: <xlsx_path>.backup.xlsx)
          2. Open xlsx; read each sidecar; append its rows
          3. Save to <xlsx_path>.tmp; on success rename tmp -> main, delete
             backup + sidecars; on failure discard tmp, leave backup +
             sidecars for retry.

        Returns: {"merged_sidecars": int, "rows_appended": int,
                  "runs_appended": int, "backup_path": str|None}

    compute_run_n(xlsx_path, *, session_id, subject, rubric_version, shape,
                  reasoner) -> int
        Count existing matching runs on the Runs tab + 1. Returns 1 for a
        brand-new workbook or unseen (session, config) combo.

The schema lives in two constants at module top so callers can introspect.

Recovery rules of thumb (PLAN.md §3.3):
  - No <xlsx>.backup.xlsx exists → last merge OK, file is clean
  - Backup exists, no merge running → last merge crashed; cp backup → main,
    re-run merge to fold in queued sidecars
  - Both dead → rebuild from per-run 5_answers.json files
"""

from __future__ import annotations

import json
import logging
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook

from pipeline.types import Rubric, RubricAnswer, RubricAnswerSet

log = logging.getLogger(__name__)


# ─── Schemas ──────────────────────────────────────────────────────────────

# Subject tabs (24 cols)
SUBJECT_COLUMNS: list[str] = [
    "run_id",                       # 1
    "session_id",                   # 2
    "session_date",                 # 3
    "camera",                       # 4
    "teacher_id",                   # 5
    "subject",                      # 6
    "rubric_version",               # 7
    "vision_model",                 # 8
    "vision_fps",                   # 9
    "chunking",                     # 10
    "shape",                        # 11
    "reasoner",                     # 12
    "run_n",                        # 13
    "question_id",                  # 14
    "section",                      # 15
    "question_text",                # 16
    "answer",                       # 17
    "confidence",                   # 18
    "evidence_timestamps",          # 19  (comma-joined HH:MM:SS list)
    "rationale",                    # 20
    "insufficient_information",     # 21
    "had_evidence",                 # 22
    "evidence_parse_ok",            # 23
    "prompt_hash",                  # 24
    # Added 2026-06-10 for typed answers (scored_1_4 / yes_no / numeric /
    # multi_choice / free_text). Appended at the end so the column order
    # of existing seeded rows stays stable.
    "answer_type",                  # 25 — from the rubric workbook col F
    "answer_type_valid",            # 26 — True if `answer` matches answer_type
]

# Runs tab (17 cols — index of every run_rubric.py invocation)
RUNS_COLUMNS: list[str] = [
    "run_id",                       # 1
    "session_id",                   # 2
    "subject",                      # 3
    "config_slug",                  # 4
    "rubric_version",               # 5
    "vision_model",                 # 6
    "shape",                        # 7
    "reasoner",                     # 8
    "run_n",                        # 9
    "started_at",                   # 10
    "finished_at",                  # 11
    "wall_clock_seconds",           # 12
    "questions_answered",           # 13
    "questions_insufficient",       # 14
    "prompt_hash",                  # 15
    "cost_usd_estimate",            # 16  (None for now; future)
    "evidence_bundle_path",         # 17  (str|None — Shape A leaves blank)
]

SUBJECT_TABS = ["Art", "Public Speaking", "Robotics"]
TAB_NAME_FROM_SUBJECT = {
    "art": "Art",
    "public_speaking": "Public Speaking",
    "robotics": "Robotics",
}

README_LINES: list[str] = [
    "TQM Answer Accumulator — README",
    "",
    "This workbook is appended-to by every scripts/run_rubric.py invocation.",
    "Each subject tab (Art / Public Speaking / Robotics) holds one row per",
    "(run × question). Re-running the same config on the same session",
    "appends new rows with run_n=2, 3, ... — full history preserved so we",
    "can compute run-to-run consistency.",
    "",
    "Tabs:",
    "  Art / Public Speaking / Robotics  — answers, 24 cols each",
    "  Runs                              — one row per run_rubric.py call",
    "                                       (audit log), 17 cols",
    "  README                            — this text",
    "",
    "Write protocol:",
    "  scripts/run_rubric.py writes a per-run sidecar JSON to",
    "  data/_answer_queue/<run_id>__<config>.json, then calls a merge that:",
    "    - copies tqm_answers.xlsx -> tqm_answers.backup.xlsx",
    "    - folds the sidecar rows into the appropriate tabs",
    "    - on success: rename tmp -> main, DELETE backup + sidecar",
    "    - on failure: discard tmp, LEAVE backup + sidecar for retry",
    "",
    "Recovery:",
    "  Backup file exists when no merge is running -> last merge crashed.",
    "  cp tqm_answers.backup.xlsx -> tqm_answers.xlsx, then re-run merge",
    "  to fold any queued sidecars.",
    "",
    "  Total disaster: rebuild from data/rubric_runs/**/5_answers.json.",
    "",
    "See PLAN.md §3.3 for the full design.",
]


# ─── Init ─────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_HEADER_FONT = Font(bold=True)


def init_workbook(xlsx_path: Path) -> bool:
    """Create a fresh empty accumulator workbook with all 5 tabs + headers.
    Idempotent — returns False if the file already exists."""
    if xlsx_path.exists():
        return False
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # First sheet is the default empty one — rename it to the first subject.
    default = wb.active
    default.title = SUBJECT_TABS[0]
    _write_header(default, SUBJECT_COLUMNS)

    for name in SUBJECT_TABS[1:]:
        ws = wb.create_sheet(name)
        _write_header(ws, SUBJECT_COLUMNS)

    runs = wb.create_sheet("Runs")
    _write_header(runs, RUNS_COLUMNS)

    readme = wb.create_sheet("README")
    for i, line in enumerate(README_LINES, start=1):
        readme.cell(row=i, column=1, value=line)
    readme.column_dimensions["A"].width = 80

    wb.save(xlsx_path)
    log.info(f"initialized empty accumulator at {xlsx_path}")
    return True


def _write_header(ws, columns: list[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


# ─── run_n computation ────────────────────────────────────────────────────


def compute_run_n(
    xlsx_path: Path,
    *,
    session_id: str,
    subject: str,
    rubric_version: str,
    shape: str,
    reasoner: str,
) -> int:
    """How many prior runs already exist for this (session × config) combo,
    plus 1. Returns 1 if the workbook doesn't exist yet."""
    if not xlsx_path.exists():
        return 1
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    if "Runs" not in wb.sheetnames:
        return 1
    runs = wb["Runs"]
    headers = [c.value for c in runs[1]]
    idx = {n: headers.index(n) for n in RUNS_COLUMNS if n in headers}
    count = 0
    for row in runs.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        if (
            row[idx["session_id"]] == session_id
            and row[idx["subject"]] == subject
            and row[idx["rubric_version"]] == rubric_version
            and row[idx["shape"]] == shape
            and row[idx["reasoner"]] == reasoner
        ):
            count += 1
    return count + 1


# ─── Sidecar emission ────────────────────────────────────────────────────


def write_sidecar(
    queue_dir: Path,
    *,
    answer_set: RubricAnswerSet,
    rubric: Rubric,
    config: dict,
    run_id: str,
    started_at: str,
    finished_at: str,
    wall_clock_seconds: float,
    config_slug: str,
    run_n: int,
) -> Path:
    """Build + write a self-contained sidecar JSON to the queue.

    Sidecar shape:
      {
        "run_id": "...",
        "subject": "art",
        "config": {...},                  ← echo of 0_config.json for audit
        "subject_rows": [{...}, ...],     ← 31 rows for the subject tab
        "runs_row": {...},                ← 1 row for the Runs tab
      }
    """
    queue_dir.mkdir(parents=True, exist_ok=True)

    # Resolve question text + answer_type from the rubric inline per row
    q_lookup = {q.id: q for q in rubric.all_questions()}

    subject_rows: list[dict[str, Any]] = []
    for qid, ans in answer_set.answers.items():
        q = q_lookup.get(qid)
        text = q.observe_text if q else "(unknown)"
        section = q.section if q else "(unknown)"
        answer_type = q.answer_type if q else "free_text"
        subject_rows.append({
            "run_id": run_id,
            "session_id": answer_set.session_id,
            "session_date": _session_date_from_id(answer_set.session_id),
            "camera": _camera_from_session_id(answer_set.session_id),
            "teacher_id": config.get("teacher_id"),  # CLI override now (option-c); auto from teacher_schedule once #33 lands
            "subject": answer_set.subject,
            "rubric_version": answer_set.rubric_version,
            "vision_model": config.get("vision_model"),
            "vision_fps": config.get("vision_fps"),
            "chunking": config.get("chunking"),
            "shape": answer_set.shape,
            "reasoner": answer_set.source_model,
            "run_n": run_n,
            "question_id": qid,
            "section": section,
            "question_text": text,
            "answer": ans.answer,
            "confidence": ans.confidence,
            "evidence_timestamps": ", ".join(ans.evidence_timestamps),
            "rationale": ans.rationale,
            "insufficient_information": ans.insufficient_information,
            "had_evidence": ans.had_evidence,
            "evidence_parse_ok": ans.evidence_parse_ok,
            "prompt_hash": answer_set.prompt_hash,
            "answer_type": answer_type,
            "answer_type_valid": ans.answer_type_valid,
        })

    runs_row = {
        "run_id": run_id,
        "session_id": answer_set.session_id,
        "subject": answer_set.subject,
        "config_slug": config_slug,
        "rubric_version": answer_set.rubric_version,
        "vision_model": config.get("vision_model"),
        "shape": answer_set.shape,
        "reasoner": answer_set.source_model,
        "run_n": run_n,
        "started_at": started_at,
        "finished_at": finished_at,
        "wall_clock_seconds": wall_clock_seconds,
        "questions_answered": sum(
            1 for a in answer_set.answers.values() if not a.insufficient_information
        ),
        "questions_insufficient": sum(
            1 for a in answer_set.answers.values() if a.insufficient_information
        ),
        "prompt_hash": answer_set.prompt_hash,
        "cost_usd_estimate": None,
        "evidence_bundle_path": config.get("evidence_bundle_path"),
    }

    sidecar = {
        "run_id": run_id,
        "subject": answer_set.subject,
        "config": config,
        "subject_rows": subject_rows,
        "runs_row": runs_row,
    }
    out = queue_dir / f"{run_id}__{config_slug}.json"
    out.write_text(json.dumps(sidecar, indent=2, default=str))
    log.info(f"wrote sidecar {out.name} — {len(subject_rows)} subject rows + 1 runs row")
    return out


def _session_date_from_id(session_id: str) -> str:
    """'2026-05-18__D28__0900' → '2026-05-18'. Returns the input on parse fail."""
    try:
        return session_id.split("__", 1)[0]
    except Exception:
        return session_id


def _camera_from_session_id(session_id: str) -> str:
    """'2026-05-18__D28__0900' → 'D28'."""
    try:
        return session_id.split("__")[1]
    except Exception:
        return ""


# ─── Atomic merge ────────────────────────────────────────────────────────


def merge_queue(
    xlsx_path: Path,
    queue_dir: Path,
    *,
    backup_path: Optional[Path] = None,
) -> dict:
    """Atomic merge: fold every sidecar in queue_dir into xlsx_path.

    Protocol (PLAN.md §3.3):
      1. Backup current xlsx to <backup_path> (default: <xlsx_path>.backup.xlsx)
      2. Open xlsx (init if missing); for each sidecar, append rows to the
         right subject tab + the Runs tab
      3. Save to <xlsx_path>.tmp
      4. On success: rename tmp -> main; delete backup + sidecars
      5. On failure: discard tmp; LEAVE backup + sidecars for retry

    Returns a dict summarising the merge.
    """
    if backup_path is None:
        backup_path = xlsx_path.with_suffix(".backup.xlsx")
    tmp_path = xlsx_path.with_suffix(".tmp.xlsx")

    sidecars = sorted(queue_dir.glob("*.json")) if queue_dir.exists() else []
    if not sidecars:
        log.info(f"no sidecars in {queue_dir} — nothing to merge")
        return {
            "merged_sidecars": 0, "rows_appended": 0,
            "runs_appended": 0, "backup_path": None,
        }

    # 1. Backup
    if xlsx_path.exists():
        shutil.copy2(xlsx_path, backup_path)
    else:
        init_workbook(xlsx_path)
        shutil.copy2(xlsx_path, backup_path)

    rows_appended = 0
    runs_appended = 0
    try:
        wb = openpyxl.load_workbook(xlsx_path)

        for sidecar_path in sidecars:
            sidecar = json.loads(sidecar_path.read_text())
            tab_name = TAB_NAME_FROM_SUBJECT.get(sidecar["subject"])
            if tab_name is None or tab_name not in wb.sheetnames:
                raise ValueError(
                    f"unknown subject {sidecar['subject']!r} in {sidecar_path.name}"
                )
            subj_ws = wb[tab_name]
            for row_dict in sidecar["subject_rows"]:
                subj_ws.append([row_dict.get(c) for c in SUBJECT_COLUMNS])
                rows_appended += 1
            runs_ws = wb["Runs"]
            runs_ws.append([sidecar["runs_row"].get(c) for c in RUNS_COLUMNS])
            runs_appended += 1

        wb.save(tmp_path)
        # 4. Success: atomic rename + cleanup
        tmp_path.replace(xlsx_path)
        for sidecar_path in sidecars:
            sidecar_path.unlink()
        backup_path.unlink(missing_ok=True)
        log.info(
            f"merge OK: {len(sidecars)} sidecars → "
            f"{rows_appended} subject rows + {runs_appended} runs rows"
        )
        return {
            "merged_sidecars": len(sidecars),
            "rows_appended": rows_appended,
            "runs_appended": runs_appended,
            "backup_path": None,
        }

    except Exception as e:
        # 5. Failure: leave sidecars + backup for next attempt
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)
        log.error(
            f"merge FAILED: {type(e).__name__}: {e}. "
            f"Backup preserved at {backup_path}; "
            f"sidecars retained in {queue_dir}."
        )
        return {
            "merged_sidecars": 0,
            "rows_appended": rows_appended,
            "runs_appended": runs_appended,
            "backup_path": str(backup_path),
            "error": f"{type(e).__name__}: {e}",
        }
