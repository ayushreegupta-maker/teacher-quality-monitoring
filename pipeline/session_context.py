"""
Turn a session_id into the session metadata + raw segment paths the
rubric pipeline needs.

`resolve_session_segments(session_id, ...)` — given a session_id like
`2026-05-18__D28__0900`, locates the subject-bucketed raw .mp4 segments
that span the class window. Subject is derived from `cctv_cameras.xlsx`
(1 camera = 1 subject). Returns SegmentEntry per matching file, sorted
chronologically. Used by `scripts/run_rubric.py`.

Pure-Python — no LLM, no ffmpeg, no Gemini calls.

The legacy DB-backed `resolve_session_context(video_path, ...)` and its
filename parser + db helpers moved to
`pipeline/_archive/session_context_legacy.py` on 2026-06-10 once they
had zero live callers. See DECISIONS.md.
"""

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Optional

import openpyxl

log = logging.getLogger(__name__)

# ─── Module roots ──────────────────────────────────────────────────────────
# Project root is two parents above this file (pipeline/session_context.py).
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_DEFAULT_RAW_DIR = _PROJECT_ROOT / "data" / "raw"
_DEFAULT_CAMERAS_XLSX = _PROJECT_ROOT / "data" / "cctv_cameras.xlsx"


# ─── Session-segment resolution (subject-bucketed raw/) ────────────────────

# Session-id format: <YYYY-MM-DD>__<camera>__<HHMM>
#   e.g. 2026-05-18__D28__0900 = D28 at 09:00 on 2026-05-18
_SESSION_ID_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})__(D\d+)__(\d{2})(\d{2})$")

# Subject-bucketed segment filename:
#   D<N>_<centre>_<subject>_<YYYYMMDD>_<HHMMSS>.mp4
# Subject can contain underscores ('public_speaking'), so we anchor on the
# trailing 8-digit date + 6-digit time tokens.
_SEGMENT_RE = re.compile(
    r"^(?P<cam>D\d+)"
    r"_(?P<centre>[a-z]+)"
    r"_(?P<subject>.+)"
    r"_(?P<date>\d{8})"
    r"_(?P<time>\d{6})"
    r"\.mp4$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class SegmentEntry:
    """One raw .mp4 segment that belongs to a session."""
    path: Path
    camera_id: str
    centre: str
    subject: str
    starts_at: datetime  # parsed from the filename's date + time suffix


def parse_session_id(session_id: str) -> tuple[date, str, time]:
    """Split a session_id like '2026-05-18__D28__0900' into (date, camera, time).

    Raises ValueError on malformed input.
    """
    m = _SESSION_ID_RE.match(session_id)
    if not m:
        raise ValueError(
            f"malformed session_id {session_id!r} — "
            f"expected <YYYY-MM-DD>__<camera>__<HHMM>"
        )
    d = date.fromisoformat(m.group(1))
    cam = m.group(2)
    t = time(int(m.group(3)), int(m.group(4)))
    return d, cam, t


def _load_camera_lookup(cameras_xlsx: Path) -> dict[str, dict[str, str]]:
    """Read cctv_cameras.xlsx 'cameras' sheet into
    {camera_id: {subject, centre_name, is_active}}.

    Camera ids are case-preserved as stored in the sheet (typically uppercase
    like D14/D28/D29).
    """
    if not cameras_xlsx.exists():
        raise FileNotFoundError(f"camera config not found: {cameras_xlsx}")
    wb = openpyxl.load_workbook(cameras_xlsx, data_only=True)
    ws = wb["cameras"]
    headers = [c.value for c in ws[1]]
    needed = {"camera_id", "centre_name", "subject", "is_active"}
    missing = needed - set(headers)
    if missing:
        raise ValueError(
            f"cctv_cameras.xlsx 'cameras' sheet missing columns: {missing}"
        )
    out: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        cam_id = rec.get("camera_id")
        if not cam_id:
            continue
        out[str(cam_id).strip()] = {
            "subject": str(rec.get("subject") or "").strip().lower(),
            "centre_name": str(rec.get("centre_name") or "").strip(),
            "is_active": bool(rec.get("is_active")),
        }
    return out


def resolve_session_segments(
    session_id: str,
    expected_duration_minutes: int = 120,
    lookback_minutes: int = 60,
    trailing_buffer_minutes: int = 30,
    raw_dir: Path = _DEFAULT_RAW_DIR,
    cameras_xlsx: Path = _DEFAULT_CAMERAS_XLSX,
) -> list[SegmentEntry]:
    """Locate the raw .mp4 segments that span (or possibly span) the class
    window for `session_id`.

    Discovery:
      1. Parse (date, camera_id, start_time) from session_id.
      2. Look up the camera's subject via cctv_cameras.xlsx (1 camera =
         1 subject is an Openhouse-wide invariant).
      3. Walk `raw_dir/<subject>/` for files matching the segment naming
         pattern with this camera and this date.
      4. Keep segments whose filename-encoded start time falls inside
            [session_start − lookback, session_start + expected_duration
             + trailing_buffer]
         The default lookback (60 min) is wide enough to capture a segment
         that began before the class but contains the start, since NVR
         chunks can be up to ~60 min long.
      5. Return entries sorted by `starts_at`.

    Returns an empty list when no segments match — caller decides whether
    that's an error or a "session not yet recorded" no-op.
    """
    session_date, camera_id, start_time = parse_session_id(session_id)
    cameras = _load_camera_lookup(cameras_xlsx)
    cam_row = cameras.get(camera_id)
    if cam_row is None:
        raise KeyError(
            f"camera_id {camera_id!r} not found in {cameras_xlsx.name}"
        )
    subject = cam_row["subject"]
    if not subject:
        raise ValueError(
            f"camera {camera_id} has no `subject` in {cameras_xlsx.name}"
        )

    subject_dir = raw_dir / subject
    if not subject_dir.exists():
        log.info(
            f"[{session_id}] no raw segments dir at {subject_dir} — empty result"
        )
        return []

    session_start_dt = datetime.combine(session_date, start_time)
    window_start = session_start_dt - timedelta(minutes=lookback_minutes)
    window_end = session_start_dt + timedelta(
        minutes=expected_duration_minutes + trailing_buffer_minutes
    )
    date_token = session_date.strftime("%Y%m%d")

    out: list[SegmentEntry] = []
    for f in subject_dir.iterdir():
        if not f.is_file() or not f.name.endswith(".mp4"):
            continue
        m = _SEGMENT_RE.match(f.name)
        if not m:
            continue
        if m.group("cam") != camera_id:
            continue
        if m.group("date") != date_token:
            continue
        # Filename subject must match camera's configured subject — sanity
        # check that the raw-bucket-by-subject migration didn't get crossed.
        if m.group("subject").lower() != subject:
            log.warning(
                f"[{session_id}] skipping {f.name}: filename subject "
                f"{m.group('subject')!r} != camera config {subject!r}"
            )
            continue
        try:
            starts_at = datetime.strptime(
                m.group("date") + m.group("time"), "%Y%m%d%H%M%S"
            )
        except ValueError:
            log.warning(f"[{session_id}] skipping {f.name}: unparseable timestamp")
            continue
        if not (window_start <= starts_at <= window_end):
            continue
        out.append(
            SegmentEntry(
                path=f,
                camera_id=camera_id,
                centre=m.group("centre"),
                subject=subject,
                starts_at=starts_at,
            )
        )

    out.sort(key=lambda e: e.starts_at)
    log.info(
        f"[{session_id}] resolved {len(out)} segment(s) in {subject_dir} "
        f"(window {window_start} → {window_end})"
    )
    return out
