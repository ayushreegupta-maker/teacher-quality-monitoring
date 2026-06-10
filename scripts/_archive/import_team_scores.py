"""
Import team rater scores from `~/Downloads/Setting Review.xlsx` into the
`team_scores` table.

The XLSX has 4 sheets. We use the two CLEAN consolidated sheets:
- "Playground Design": one row per (dimension, video_number), with rater scores
- "Toy Design": same shape for the toy_design rubric

Layout for both sheets:
  Col 0: dimension name
  Col 1: video number (1.0–7.0)
  Cols 2–5/6: "Setting 1/2/3/4" (different model attempts — we ignore these)
  Cols 6–8 (PG) or 5–7 (TD): Akshay, Pari, Ayesha desired scores

Video-number → manifest-file mapping is inferred from the comments in Sheet1
(see DECISIONS.md and the analysis in import_team_scores notes). The
mapping below is the best guess; correct here if any video is misidentified.
"""

import sys
from pathlib import Path

# Ensure project root on sys.path
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

import adapters.db as db

XLSX_PATH = Path.home() / "Downloads" / "Setting Review.xlsx"

# Map team video number → manifest video filename (without .MOV extension lowercased).
# Authoritative mapping confirmed by Openhouse on 2026-05-26. This also
# resolves the "Open mapping question (Video 2)" from DECISIONS.md — Video 2
# really is multi-sensory bins (the rater comments about water/fishing
# were a different rater's confusion, not a different video).
VIDEO_NUMBER_TO_FILE = {
    1: "floorpainting",
    2: "multisensorybins_1505",
    3: "damru_1",
    4: "naturesensorybin_1",
    5: "sandplay_1",
    6: "naturesensorybin_hrbr",
    7: "waterplay",
}

# Map XLSX dimension labels → our rubric dimension IDs.
DIMENSION_NAME_MAP = {
    # Playground Design (rubric_name = "playground", rubric_version = "0.2.3")
    "Narrative Setting": ("playground", "narrative"),
    "Multi Sensory": ("playground", "multi_sensory"),
    "Boundary": ("playground", "boundary"),
    "Movement & Layout": ("playground", "movement"),
    "Cleanup & Resetting": ("playground", "clean_up"),
    # Toy Design (rubric_name = "toy_design", rubric_version = "0.2.0")
    "Purpose": ("toy_design", "purpose"),
    "Anchor & Choice Materials": ("toy_design", "anchor_and_choice_materials"),
    "Spark Curiosity": ("toy_design", "spark_curiosity"),
    "Challenge Adjustment": ("toy_design", "challenge_adjustment"),
    "Self Served": ("toy_design", "self_served"),
}

# Rubric version stamps. Match the YAML rubric files.
RUBRIC_VERSIONS = {
    "playground": "0.2.3",
    "toy_design": "0.2.0",
}

RATERS = ["Akshay", "Pari", "Ayesha"]


def _session_id_for(video_file_stem: str) -> str:
    """We use the same convention as batch_score.py: `batch_<rubric>_<slug>`.
    But team scores are per-video, not per-rubric run, so we use a generic
    session_id keyed on the manifest stem. The (session_id, rater, rubric, dim)
    UNIQUE constraint in team_scores then disambiguates correctly.
    """
    return f"manifest_{video_file_stem.lower()}"


def parse_sheet(ws, rubric_name: str) -> list[dict]:
    """Parse one consolidated sheet (Playground Design or Toy Design).

    Returns a list of dicts ready for db.save_team_score(): video_number,
    dimension_id, rater_name, score.
    """
    rows = []
    # Read header row to find rater column indices
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        akshay_col = header.index("Akshay Desired Score")
        pari_col = header.index("Pari Desired Score")
        ayesha_col = header.index("Ayesha Desired Score")
    except ValueError as e:
        raise RuntimeError(f"Could not find rater columns in {ws.title}: {e}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        dim_label = row[0]
        if not dim_label:
            continue
        video_num = row[1]
        if video_num is None:
            continue
        try:
            video_num_int = int(float(video_num))
        except (TypeError, ValueError):
            continue
        if video_num_int not in VIDEO_NUMBER_TO_FILE:
            continue
        rubric_dim = DIMENSION_NAME_MAP.get(dim_label)
        if not rubric_dim:
            # Skip unknown dimensions silently (e.g. "Total Playground Design")
            continue
        if rubric_dim[0] != rubric_name:
            continue  # safety: dimension belongs to the OTHER rubric
        for rater_idx, rater in zip([akshay_col, pari_col, ayesha_col], RATERS):
            score = row[rater_idx]
            if score is None or score == "":
                continue
            rows.append({
                "video_number": video_num_int,
                "dimension_id": rubric_dim[1],
                "rater_name": rater,
                "score": float(score),
            })
    return rows


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found.")
        sys.exit(1)

    # Ensure DB is initialized
    db.init_db()

    wb = load_workbook(XLSX_PATH, data_only=True)
    if "Playground Design" not in wb.sheetnames or "Toy Design" not in wb.sheetnames:
        print(f"ERROR: Expected sheets 'Playground Design' and 'Toy Design'; got {wb.sheetnames}")
        sys.exit(1)

    pg_rows = parse_sheet(wb["Playground Design"], "playground")
    td_rows = parse_sheet(wb["Toy Design"], "toy_design")
    all_rows = pg_rows + td_rows

    print(f"Parsed {len(pg_rows)} playground rows + {len(td_rows)} toy_design rows = {len(all_rows)} total")
    print()

    # Persist
    inserted = 0
    for r in all_rows:
        video_stem = VIDEO_NUMBER_TO_FILE[r["video_number"]]
        rubric_name = "playground" if r["dimension_id"] in [
            "narrative", "multi_sensory", "boundary", "movement", "clean_up"
        ] else "toy_design"
        db.save_team_score(
            session_id=_session_id_for(video_stem),
            rater_name=r["rater_name"],
            rubric_name=rubric_name,
            rubric_version=RUBRIC_VERSIONS[rubric_name],
            dimension=r["dimension_id"],
            score=r["score"],
        )
        inserted += 1

    print(f"Inserted/updated {inserted} team_score rows in DB.")
    print()

    # Quick sanity print: per-rater coverage
    with db.db_conn() as conn:
        result = conn.execute(
            "SELECT rater_name, rubric_name, COUNT(*) AS n "
            "FROM team_scores GROUP BY rater_name, rubric_name "
            "ORDER BY rater_name, rubric_name"
        ).fetchall()
        print("Per-rater coverage:")
        for row in result:
            print(f"  {row['rater_name']:8s} {row['rubric_name']:12s} {row['n']:3d} rows")

        # Per-video coverage
        print()
        result = conn.execute(
            "SELECT session_id, COUNT(*) AS n FROM team_scores "
            "GROUP BY session_id ORDER BY session_id"
        ).fetchall()
        print("Per-video row count:")
        for row in result:
            print(f"  {row['session_id']:50s} {row['n']:3d} rows")


if __name__ == "__main__":
    main()
