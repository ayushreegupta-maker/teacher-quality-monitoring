"""
Q&A rubric loader + prompt renderer.

The rubric workbook (`prompts/rubrics.xlsx`) has
one tab per subject: Art, Public Speaking, Robotics. Each tab is a flat
table of questions with one row per question and these columns:

  A. (unlabelled)         — section name, sparsely populated; carry-forward
                            until the next non-empty value. The four
                            distinct section values are
                            Environment / Content Knowledge / Facilitation / Warmth.
  B. "Criteria"           — group label, also carry-forward.
  C. "What AI needs to    — the actual question text. ONE row = ONE question.
       observe"
  D. "Input required"     — optional input reference (present in Art tab only).
  E. "Analysis"           — Visual / Audio / Visual + Audio tag. Present in
                            Art tab only; default-filled for the others
                            (PS + Robotics tabs are at v0 and don't carry
                            this column yet — see PLAN.md §2).

This module exposes ONE entry point — `load_rubric(workbook_path, subject)` —
that returns a fully typed `Rubric` (Pydantic; defined in pipeline.types).
Prompt rendering and scoring land in this module in step 8 of the migration.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl

from adapters.llm import LLMAdapter, parse_json_lenient, prompt_hash
from pipeline.types import (
    EvidenceBundle,
    MaterialSeen,
    Rubric,
    RubricAnswer,
    RubricAnswerSet,
    RubricQuestion,
    RubricSection,
)

# Default reasoner for Shape B (Q2 of step 9). Calibrated baseline from the
# one-off score_art_with_claude.py experiment. Callers can override via
# `reasoner_model=` to sweep cheaper alternatives.
DEFAULT_SHAPE_B_REASONER = "claude-opus-4-7"

log = logging.getLogger(__name__)

# Subject → workbook sheet name. The rubric workbook uses human-readable
# tab names, but the rest of the pipeline uses snake_case subject tokens
# everywhere else (cctv_cameras.xlsx, data/raw/<subject>/, etc.).
_SUBJECT_TO_SHEET = {
    "art": "Art",
    "public_speaking": "Public Speaking",
    "robotics": "Robotics",
}

# Default analysis tag for any question row that doesn't carry one (PS and
# Robotics tabs in the current workbook). Matches the most permissive option
# in the Art tab so the rubric prompt isn't artificially narrowed.
DEFAULT_ANALYSIS_TAG = "Visual + Audio"


# ─── answer_type validation + level/option parsing ────────────────────────

_VALID_ANSWER_TYPES = {"scored_1_4", "yes_no", "numeric", "multi_choice", "free_text"}


def _normalise_answer_type(raw: str | None, qid: str, subject: str) -> str:
    """Read column F and return the canonical answer_type. Blank → 'free_text'.
    Unknown values warn + fall back to 'free_text'."""
    if raw is None:
        return "free_text"
    key = raw.strip()
    if key in _VALID_ANSWER_TYPES:
        return key
    log.warning(
        f"[{subject}] {qid}: unknown answer_type {raw!r}; falling back to "
        f"'free_text'. Valid: {sorted(_VALID_ANSWER_TYPES)}"
    )
    return "free_text"


def _parse_levels(
    h: str | None, i: str | None, j: str | None, k: str | None,
    answer_type: str, qid: str, subject: str,
) -> list[str] | None:
    """For scored_1_4: return the [lvl1, lvl2, lvl3, lvl4] list. Warns + returns
    None if not all 4 are present (downstream renderer treats as free_text).
    For other types: returns None (levels columns ignored)."""
    if answer_type != "scored_1_4":
        return None
    raw = [h, i, j, k]
    filled = [x for x in raw if x]
    if len(filled) != 4:
        log.warning(
            f"[{subject}] {qid}: answer_type='scored_1_4' needs 4 non-empty "
            f"level cells (cols H-K); got {len(filled)}. Levels not used."
        )
        return None
    return [str(x) for x in raw]


def load_rubric(
    workbook_path: Path,
    subject: str,
    default_analysis_tag: str = DEFAULT_ANALYSIS_TAG,
) -> Rubric:
    """Load one subject's tab from the rubric workbook into a typed Rubric.

    Behaviour:
      - Iterates rows from row 2 onward.
      - Col A (section) and col B (criteria) are sparsely populated; the
        loader forward-fills from the most recent non-empty value so every
        question carries the section + criteria it belongs to.
      - A question row is any row whose col C (observe_text) is non-empty.
      - Question ids are assigned in encounter order: "Q1", "Q2", ...
      - Sections are emitted in encounter order (first time a new section
        name appears, a new RubricSection is opened).
      - Missing col D (input_ref) → None.
      - Missing col E (analysis_tag) → `default_analysis_tag` (PS+Robotics).

    Raises:
      FileNotFoundError if `workbook_path` doesn't exist.
      ValueError if `subject` doesn't map to a sheet in the workbook.
      ValueError if the sheet has no question rows.
    """
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"rubric workbook not found: {workbook_path}")

    sheet_name = _SUBJECT_TO_SHEET.get(subject)
    if sheet_name is None:
        raise ValueError(
            f"unknown subject {subject!r} — "
            f"expected one of {sorted(_SUBJECT_TO_SHEET)}"
        )

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"sheet {sheet_name!r} not in workbook (have: {wb.sheetnames})"
        )
    ws = wb[sheet_name]

    sections: list[RubricSection] = []
    # Maintain an open section we keep appending questions to until the
    # section name in col A changes.
    open_section: RubricSection | None = None
    current_section: str | None = None
    current_criteria: str | None = None
    q_counter = 0

    # Workbook column layout (same across all tabs; PS + Robotics simply
    # leave D-H blank, which falls through to free_text):
    #   A: section / bucket   (carry-forward)
    #   B: criteria           (carry-forward)
    #   C: question text      — required
    #   D: answer_type        — one of the 5 canonical tokens
    #   E: level_1            — scored_1_4 only
    #   F: level_2
    #   G: level_3
    #   H: level_4

    def _cell(row, idx):
        if idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    for row in ws.iter_rows(min_row=2, values_only=True):
        col_a = _cell(row, 0)   # section
        col_b = _cell(row, 1)   # criteria
        col_c = _cell(row, 2)   # question
        col_d = _cell(row, 3)   # answer_type
        col_e = _cell(row, 4)   # level_1
        col_f = _cell(row, 5)   # level_2
        col_g = _cell(row, 6)   # level_3
        col_h = _cell(row, 7)   # level_4

        if col_a:
            current_section = col_a
        if col_b:
            current_criteria = col_b

        if col_c is None:
            continue

        if current_section is None:
            log.warning(
                f"[{subject}] question row with no section above it at "
                f"row {q_counter + 2}, skipping: {col_c[:60]}"
            )
            continue

        if open_section is None or open_section.name != current_section:
            open_section = RubricSection(name=current_section, questions=[])
            sections.append(open_section)

        q_counter += 1
        qid = f"Q{q_counter}"

        answer_type = _normalise_answer_type(col_d, qid, subject)
        levels = _parse_levels(col_e, col_f, col_g, col_h,
                               answer_type, qid, subject)

        question = RubricQuestion(
            id=qid,
            section=current_section,
            criteria=current_criteria,
            observe_text=col_c,
            analysis_tag=default_analysis_tag,
            answer_type=answer_type,
            levels=levels,
        )
        open_section.questions.append(question)

    if q_counter == 0:
        raise ValueError(
            f"sheet {sheet_name!r} contained no question rows "
            "(no non-empty 'What AI needs to observe' cells)"
        )

    rubric = Rubric(
        subject=subject,
        source_path=str(workbook_path.resolve()),
        sections=sections,
    )
    log.info(
        f"[{subject}] loaded rubric: {q_counter} question(s) across "
        f"{len(sections)} section(s) from {workbook_path.name}"
    )
    return rubric


# ─── Prompt rendering ─────────────────────────────────────────────────────


def render_questions_block(rubric: Rubric) -> str:
    """Render the rubric questions as the `{questions_block}` substring
    that gets interpolated into the prompt template.

    Output structure:

        \n=== <section> ===
        \n[Group] <criteria>
          Q1 [<analysis_tag>] (<answer_type_hint>): <observe_text>  (input ref: ...)
              Description: <description>            ← if present
              1 = <level_1_worst>                   ← scored_1_4 only
              2 = <level_2>
              3 = <level_3>
              4 = <level_4_best>
              Options: a, b, c                       ← multi_choice only

    Backwards-compatible: questions with answer_type='free_text' and no
    description / levels / options render exactly like before (no type hint,
    no description line).
    """
    lines: list[str] = []
    current_section: str | None = None
    current_criteria: str | None = None
    for q in rubric.all_questions():
        if q.section != current_section:
            current_section = q.section
            current_criteria = None
            lines.append(f"\n=== {current_section} ===")
        if q.criteria != current_criteria:
            current_criteria = q.criteria
            lines.append(f"\n[Group] {current_criteria}")
        lines.extend(_render_one_question(q))
    return "\n".join(lines)


_TYPE_HINTS: dict[str, str] = {
    "scored_1_4": "scored 1-4",
    "yes_no": "yes/no",
    "numeric": "integer",
    "multi_choice": "pick from options",
    "free_text": "",  # no hint → backward-compatible rendering
}


def _render_one_question(q: RubricQuestion) -> list[str]:
    """Render a single RubricQuestion as a list of lines."""
    analysis_tag = f"[{q.analysis_tag}]" if q.analysis_tag else ""
    type_hint = _TYPE_HINTS.get(q.answer_type, "")
    type_hint_str = f" ({type_hint})" if type_hint else ""
    input_hint = f"  (input ref: {q.input_ref})" if q.input_ref else ""

    head = f"  {q.id} {analysis_tag}{type_hint_str}: {q.observe_text}{input_hint}"
    lines = [head]

    if q.description:
        lines.append(f"      Description: {q.description}")

    if q.answer_type == "scored_1_4" and q.levels and len(q.levels) == 4:
        lines.append(f"      1 = {q.levels[0]}")
        lines.append(f"      2 = {q.levels[1]}")
        lines.append(f"      3 = {q.levels[2]}")
        lines.append(f"      4 = {q.levels[3]}")
    elif q.answer_type == "multi_choice" and q.options:
        lines.append(f"      Options: {', '.join(q.options)}")

    return lines


# Shape labels mirror PLAN.md §3 — "A" = Gemini watches the video directly,
# "B" = Gemini extracted evidence, Claude (or another text reasoner) scores.
_SHAPE_A = "A"
_SHAPE_B = "B"


def render_prompt(
    rubric: Rubric,
    prompt_path: Path,
    *,
    shape: str = _SHAPE_A,
    # Shape A metadata:
    duration_str: str | None = None,
    duration_sec: int | None = None,
    wallclock_start: str | None = None,
    wallclock_end: str | None = None,
    # Shape B evidence:
    evidence: EvidenceBundle | None = None,
) -> str:
    """Load a rubric prompt template from `prompt_path` and interpolate it
    with the rubric's question block + the run inputs.

    Shape A (Gemini direct):
      Loads e.g. prompts/art/rubric_art_v1_2026-06-10.md. Placeholders:
      `{questions_block}`, `{duration_str}`, `{duration_sec}`,
      `{wallclock_start}`, `{wallclock_end}`. Returns one string ready to
      send alongside the trimmed video.

    Shape B (text reasoner over Gemini-extracted evidence):
      Loads e.g. prompts/art/rubric_art_v1_2026-06-10_shape_b.md.
      Placeholders: `{questions_block}` + 6 JSON blobs derived from the
      EvidenceBundle (boundaries, phases, explanations, disturbances,
      observations, transcript). Returns one string with `# SYSTEM` and
      `# USER` markers that pipeline.render.split_system_user parses out
      for the Anthropic message format.

    `prompt_path` may be an absolute Path or relative (we search both
    `<rel>` and `prompts/<rel>` and `prompts/<rel>.md`).
    """
    # ── Resolve prompt path ──
    if not isinstance(prompt_path, Path):
        prompt_path = Path(prompt_path)
    if not prompt_path.is_absolute():
        candidates = [
            prompt_path,
            _PROMPTS_DIR / prompt_path,
            _PROMPTS_DIR / f"{prompt_path}.md",
        ]
        for cand in candidates:
            if cand.exists():
                prompt_path = cand
                break
        else:
            raise FileNotFoundError(
                f"prompt not found under {_PROMPTS_DIR}: tried {candidates}"
            )

    raw = prompt_path.read_text()
    body = _strip_frontmatter(raw)
    questions_block = render_questions_block(rubric)

    if shape == _SHAPE_A:
        missing = [
            n for n, v in (
                ("duration_str", duration_str),
                ("duration_sec", duration_sec),
                ("wallclock_start", wallclock_start),
                ("wallclock_end", wallclock_end),
            ) if v is None
        ]
        if missing:
            raise ValueError(
                f"render_prompt(shape='A') requires: {missing}"
            )
        return body.format(
            questions_block=questions_block,
            duration_str=duration_str,
            duration_sec=duration_sec,
            wallclock_start=wallclock_start,
            wallclock_end=wallclock_end,
        )

    if shape == _SHAPE_B:
        if evidence is None:
            raise ValueError(
                "render_prompt(shape='B') requires evidence=<EvidenceBundle>"
            )
        return body.format(
            activity_context=evidence.activity_context or "(not specified)",
            questions_block=questions_block,
            boundaries_json=json.dumps(evidence.boundaries, indent=2),
            phases_json=json.dumps(evidence.phases or [], indent=2),
            explanations_json=json.dumps(evidence.explanations or [], indent=2),
            disturbances_json=json.dumps(evidence.disturbances or [], indent=2),
            observations_json=json.dumps(evidence.observations, indent=2),
            transcript_json=json.dumps({"segments": evidence.transcript}, indent=2),
        )

    raise ValueError(f"unknown shape {shape!r}")


# ─── Internal helpers ─────────────────────────────────────────────────────

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_PROMPTS_DIR = _PROJECT_ROOT / "prompts"

import re as _re

_FRONTMATTER_RE = _re.compile(r"^---\n.*?\n---\n", _re.DOTALL)


def _strip_frontmatter(text: str) -> str:
    """Strip a leading `---\\n...\\n---\\n` YAML frontmatter block, if present.
    Mirrors pipeline.render.strip_frontmatter — duplicated here so this
    module has no dependency on the legacy render module."""
    if not text.startswith("---"):
        return text
    m = _FRONTMATTER_RE.match(text)
    return text[m.end():] if m else text


_HMS_RE = _re.compile(r"^\d{1,2}:\d{2}:\d{2}$")


def _is_valid_hms(s: str) -> bool:
    """Strict HH:MM:SS or H:MM:SS check. Does NOT validate field ranges
    (24h, 60m, 60s); just shape."""
    if not isinstance(s, str):
        return False
    return bool(_HMS_RE.match(s.strip()))


def _validate_answer_against_type(
    ans_str: str,
    question: RubricQuestion,
    is_insufficient: bool,
    session_id: str,
    qid: str,
) -> bool:
    """Return True if the answer matches the question's declared answer_type.
    INSUFFICIENT INFORMATION is always considered valid regardless of type.
    Logs a warning + returns False on type mismatch."""
    if is_insufficient:
        return True

    at = question.answer_type
    if at == "free_text":
        return True  # anything goes

    if at == "scored_1_4":
        try:
            n = int(ans_str.strip())
            if 1 <= n <= 4:
                return True
        except ValueError:
            pass
        log.warning(
            f"[{session_id}] {qid}: answer_type=scored_1_4 expected integer "
            f"1-4, got {ans_str!r}"
        )
        return False

    if at == "yes_no":
        norm = ans_str.strip().lower().rstrip(".!")
        if norm in ("yes", "no", "y", "n", "true", "false"):
            return True
        # Forgive longer answers that START with yes/no
        if norm.startswith(("yes", "no", "y ", "n ", "yes,", "no,")):
            return True
        log.warning(
            f"[{session_id}] {qid}: answer_type=yes_no expected Yes/No, "
            f"got {ans_str!r}"
        )
        return False

    if at == "numeric":
        try:
            int(ans_str.strip().rstrip("."))
            return True
        except ValueError:
            pass
        # Forgive "6 minutes" / "approximately 6" / "~6"
        m = _re.search(r"-?\d+", ans_str)
        if m is not None:
            return True
        log.warning(
            f"[{session_id}] {qid}: answer_type=numeric expected an integer, "
            f"got {ans_str!r}"
        )
        return False

    if at == "multi_choice":
        if not question.options:
            return True  # loader downgraded to free_text effectively
        opts_lower = {o.lower().strip() for o in question.options}
        # Accept comma/semicolon-separated multi-select
        chosen = _re.split(r"[,;]\s*", ans_str.lower().strip())
        chosen = [c.strip() for c in chosen if c.strip()]
        if not chosen:
            return False
        for c in chosen:
            if c not in opts_lower:
                log.warning(
                    f"[{session_id}] {qid}: multi_choice answer {c!r} not in "
                    f"declared options {question.options}"
                )
                return False
        return True

    return True  # unknown type — be permissive


# ─── Scoring ──────────────────────────────────────────────────────────────


def score(
    *,
    rubric: Rubric,
    prompt: str,
    llm: LLMAdapter,
    session_id: str,
    rubric_version: str,
    video_path: Path | None = None,
    shape: str = _SHAPE_A,
    reasoner_model: str | None = None,
    fps: float | None = None,
    media_resolution: str | None = None,
) -> tuple[RubricAnswerSet, str]:
    """Run one rubric scoring call. Returns (RubricAnswerSet, raw_response).

    Shape A (Gemini direct):
      `video_path` is required. Uploads to Gemini, calls with the rendered
      prompt. `reasoner_model` overrides the adapter's vision model.

    Shape B (text reasoner over the evidence bundle):
      `video_path` is ignored (the caller renders evidence into the prompt).
      `reasoner_model` defaults to DEFAULT_SHAPE_B_REASONER ('claude-opus-4-7').
      The prompt must contain `# SYSTEM` and `# USER` markers — they're
      split out for the Anthropic message format.

    Both shapes return RubricAnswerSet with denormalised flags
    (insufficient_information / had_evidence / evidence_parse_ok) populated
    on each answer for downstream pivoting. Raw model response is returned
    separately so the caller can persist it for audit.
    """
    if shape == _SHAPE_A:
        if video_path is None:
            raise ValueError("score(shape='A') requires video_path")
        if not video_path.exists():
            raise FileNotFoundError(f"video not found: {video_path}")
        model = reasoner_model or llm.vision_model
        log.info(
            f"[{session_id}] score (A): uploading {video_path.name} + asking "
            f"{model} about {len(rubric.all_questions())} questions"
            + (f" (fps={fps})" if fps is not None else "")
        )
        video_file = llm.upload_video(video_path)
        raw = llm.call_gemini_video(
            prompt=prompt,
            video_file=video_file,
            model_name=reasoner_model,
            fps=fps,
            media_resolution=media_resolution,
        )
        source_model = model

    elif shape == _SHAPE_B:
        # The Shape B prompt embeds # SYSTEM and # USER markers; split them
        # out for the Anthropic message format.
        from pipeline.render import split_system_user
        system_part, user_part = split_system_user(prompt)
        model = reasoner_model or DEFAULT_SHAPE_B_REASONER
        log.info(
            f"[{session_id}] score (B): asking {model} about "
            f"{len(rubric.all_questions())} questions "
            f"(prompt: {len(system_part):,} sys / {len(user_part):,} user chars)"
        )
        # Opus 4.7 rejects the temperature parameter; we pass None so the
        # adapter omits it. Older Claude models accept temperature=0 — pass
        # 0.0 there if we ever route to one.
        temperature = None if model.startswith("claude-opus-4-7") else 0.0
        raw = llm.call_claude_text(
            system=system_part,
            user=user_part,
            max_tokens=16000,
            temperature=temperature,
            model_name=model,
        )
        source_model = model

    else:
        raise ValueError(f"unknown shape {shape!r}")

    # ── Parse + build the answer set (shared path) ──
    parsed = parse_json_lenient(raw)
    if not isinstance(parsed, dict):
        raise ValueError(
            f"expected dict response keyed by Q-id, got {type(parsed).__name__}"
        )

    answer_set = _build_answer_set(
        parsed_response=parsed,
        rubric=rubric,
        session_id=session_id,
        rubric_version=rubric_version,
        shape=shape,
        source_model=source_model,
        prompt=prompt,
    )
    log.info(
        f"[{session_id}] score ({shape}): {len(answer_set.answers)} answers parsed "
        f"({sum(1 for a in answer_set.answers.values() if a.insufficient_information)} "
        "INSUFFICIENT)"
    )
    return answer_set, raw


def _build_answer_set(
    *,
    parsed_response: dict,
    rubric: Rubric,
    session_id: str,
    rubric_version: str,
    shape: str,
    source_model: str,
    prompt: str,
) -> RubricAnswerSet:
    """Walk the model's JSON response, validate per-question shape, and
    assemble a typed RubricAnswerSet. Skips unexpected keys + malformed
    rows with a warning — never raises on a single bad answer."""
    questions_by_id = {q.id: q for q in rubric.all_questions()}
    valid_ids = set(questions_by_id.keys())
    answers: dict[str, RubricAnswer] = {}

    # Top-level meta keys the reasoner is allowed to emit alongside Q*.
    # Extracted out of the main loop so they don't trigger "unexpected qid".
    materials_seen = _parse_materials_seen(
        parsed_response.pop("materials_seen", None), session_id,
    )

    for qid, payload in parsed_response.items():
        if qid not in valid_ids:
            log.warning(f"[{session_id}] unexpected qid {qid!r}, skipping")
            continue
        if not isinstance(payload, dict):
            log.warning(
                f"[{session_id}] {qid}: payload not a dict ({type(payload).__name__}), "
                "skipping"
            )
            continue

        ans_str = str(payload.get("answer", "")).strip()
        is_insufficient = ans_str.upper().startswith("INSUFFICIENT")

        ev_ts = payload.get("evidence_timestamps") or []
        if isinstance(ev_ts, str):
            ev_ts = [ev_ts]
        ev_ts = [str(t).strip() for t in ev_ts if str(t).strip()]
        had_evidence = bool(ev_ts)
        evidence_parse_ok = all(_is_valid_hms(t) for t in ev_ts) if had_evidence else True

        confidence = str(payload.get("confidence", "low")).strip().lower()
        if confidence not in ("high", "medium", "low"):
            log.warning(
                f"[{session_id}] {qid}: confidence={confidence!r} not in "
                "{high, medium, low}; coercing to 'low'"
            )
            confidence = "low"

        # Validate the answer's shape against the question's declared
        # answer_type. INSUFFICIENT is always valid regardless of type.
        question = questions_by_id[qid]
        answer_type_valid = _validate_answer_against_type(
            ans_str, question, is_insufficient, session_id, qid,
        )

        try:
            answers[qid] = RubricAnswer(
                id=qid,
                answer=ans_str,
                confidence=confidence,
                evidence_timestamps=ev_ts,
                rationale=payload.get("rationale"),
                insufficient_information=is_insufficient,
                had_evidence=had_evidence,
                evidence_parse_ok=evidence_parse_ok,
                answer_type_valid=answer_type_valid,
            )
        except Exception as e:
            log.warning(f"[{session_id}] {qid}: validation failed ({e!r}); skipping")

    missing = sorted(valid_ids - set(answers.keys()))
    if missing:
        log.warning(
            f"[{session_id}] {len(missing)} questions not in model response: "
            f"{missing[:5]}{'...' if len(missing) > 5 else ''}"
        )

    return RubricAnswerSet(
        session_id=session_id,
        subject=rubric.subject,
        rubric_version=rubric_version,
        answers=answers,
        source_model=source_model,
        shape=shape,
        prompt_hash=prompt_hash(prompt),
        materials_seen=materials_seen,
    )


def _parse_materials_seen(raw, session_id: str) -> list[MaterialSeen] | None:
    """Coerce the reasoner's `materials_seen` field into a list[MaterialSeen].
    Returns None if absent. Skips malformed entries with a warning."""
    if raw is None:
        return None
    if not isinstance(raw, list):
        log.warning(
            f"[{session_id}] materials_seen not a list ({type(raw).__name__}); "
            "ignoring"
        )
        return None
    out: list[MaterialSeen] = []
    for i, m in enumerate(raw):
        if not isinstance(m, dict):
            log.warning(
                f"[{session_id}] materials_seen[{i}] not a dict "
                f"({type(m).__name__}); skipping"
            )
            continue
        try:
            out.append(MaterialSeen(**m))
        except Exception as e:
            log.warning(
                f"[{session_id}] materials_seen[{i}] validation failed "
                f"({e!r}); skipping"
            )
    return out
